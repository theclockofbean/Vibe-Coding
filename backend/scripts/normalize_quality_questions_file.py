# ruff: noqa: E402,I001
"""Normalize quality_questions.xlsx for Phase 3-I-B ingestion."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Final

from openpyxl import load_workbook  # type: ignore[import-untyped]


BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

QUALITY_FILE: Final[Path] = (
    PROJECT_ROOT
    / "data"
    / "uploads"
    / "qa_pairs"
    / "quality_questions.xlsx"
)

QA_SHEET_NAME: Final[str] = "qa_pairs"


def normalize_quality_questions_file() -> None:
    """Normalize quality_questions.xlsx in place with backup."""

    if not QUALITY_FILE.exists():
        raise FileNotFoundError(f"quality_questions.xlsx not found: {QUALITY_FILE}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = QUALITY_FILE.with_name(
        f"{QUALITY_FILE.stem}.backup_{timestamp}{QUALITY_FILE.suffix}"
    )
    backup_file.write_bytes(QUALITY_FILE.read_bytes())

    workbook = load_workbook(QUALITY_FILE)

    if QA_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"sheet not found: {QA_SHEET_NAME}")

    sheet = workbook[QA_SHEET_NAME]

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]

    header_to_col = {
        header: index + 1
        for index, header in enumerate(headers)
        if header
    }

    expected_source_col = ensure_column(
        sheet=sheet,
        header_to_col=header_to_col,
        header="expected_source",
    )
    notes_col = ensure_column(
        sheet=sheet,
        header_to_col=header_to_col,
        header="notes",
    )
    metadata_col = ensure_column(
        sheet=sheet,
        header_to_col=header_to_col,
        header="metadata",
    )

    answer_source_col = header_to_col.get("answer_source")
    review_notes_col = header_to_col.get("review_notes")
    source_group_id_col = header_to_col.get("source_group_id")
    risk_flags_col = header_to_col.get("risk_flags")
    qa_id_col = header_to_col.get("qa_id")

    if source_group_id_col is None:
        raise RuntimeError("source_group_id column not found")

    if qa_id_col is None:
        raise RuntimeError("qa_id column not found")

    for row_index in range(2, sheet.max_row + 1):
        qa_id = text(sheet.cell(row=row_index, column=qa_id_col).value)

        if not qa_id:
            continue

        if answer_source_col is not None:
            expected_source = text(
                sheet.cell(row=row_index, column=expected_source_col).value
            )

            if not expected_source:
                sheet.cell(
                    row=row_index,
                    column=expected_source_col,
                    value=text(sheet.cell(row=row_index, column=answer_source_col).value),
                )

        if review_notes_col is not None:
            notes = text(sheet.cell(row=row_index, column=notes_col).value)

            if not notes:
                sheet.cell(
                    row=row_index,
                    column=notes_col,
                    value=text(sheet.cell(row=row_index, column=review_notes_col).value),
                )

        metadata = text(sheet.cell(row=row_index, column=metadata_col).value)

        if not metadata:
            sheet.cell(row=row_index, column=metadata_col, value="{}")

        source_group_id = text(
            sheet.cell(row=row_index, column=source_group_id_col).value
        )

        if not source_group_id.startswith("QUAL_SRC"):
            number = extract_number(qa_id=qa_id, fallback=row_index - 1)
            sheet.cell(
                row=row_index,
                column=source_group_id_col,
                value=f"QUAL_SRC{number:04d}",
            )

        if risk_flags_col is not None:
            risk_flags = text(sheet.cell(row=row_index, column=risk_flags_col).value)
            normalized = risk_flags.replace(
                "aftersales_dispute",
                "aftersale_commitment",
            )
            sheet.cell(row=row_index, column=risk_flags_col, value=normalized)

    workbook.save(QUALITY_FILE)

    print("normalized quality_questions.xlsx")
    print(f"file={QUALITY_FILE}")
    print(f"backup={backup_file}")


def ensure_column(
    *,
    sheet,
    header_to_col: dict[str, int],
    header: str,
) -> int:
    """Ensure header exists and return column index."""

    if header in header_to_col:
        return header_to_col[header]

    new_col = sheet.max_column + 1
    sheet.cell(row=1, column=new_col, value=header)
    header_to_col[header] = new_col

    return new_col


def extract_number(
    *,
    qa_id: str,
    fallback: int,
) -> int:
    """Extract number from QUAL0001 style ID."""

    digits = "".join(char for char in qa_id if char.isdigit())

    if not digits:
        return fallback

    return int(digits)


def text(value: object) -> str:
    """Return stripped text."""

    if value is None:
        return ""

    return str(value).strip()


def main() -> int:
    """Run normalizer."""

    normalize_quality_questions_file()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())