# ruff: noqa: E402,I001
"""Inspect intent classifier behavior for Phase 3-I-I P0 routing failures.

This script intentionally does not import app.agent.parsers.* because the current
parser package has circular imports through handlers.
"""

from __future__ import annotations

import sys
from pathlib import Path
from pprint import pprint
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]

BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.agent.llm.intent_classifier import classify_intent_by_keywords


TEST_CASES_FILE: Final[Path] = PROJECT_ROOT / "data/evaluation/test_cases_draft.xlsx"
INTENT_CLASSIFIER_FILE: Final[Path] = (
    BACKEND_ROOT / "app/agent/llm/intent_classifier.py"
)

P0_CASE_IDS: Final[tuple[str, ...]] = (
    "TC_SPEC_004",
    "TC_SPEC_007",
    "TC_SPEC_011",
    "TC_LOGI_002",
    "TC_LOGI_003",
    "TC_PRICE_002",
    "TC_PRICE_007",
    "TC_PRICE_008",
    "TC_QUAL_001",
    "TC_QUAL_003",
    "TC_QUAL_004",
    "TC_QUAL_005",
    "TC_QUAL_008",
    "TC_ESCA_001",
    "TC_ESCA_002",
    "TC_ESCA_003",
)

SOURCE_ANCHORS: Final[tuple[str, ...]] = (
    "def classify_intent_by_keywords",
    "def _resolve_llm_intent_with_local_cues",
    "ALLOWED_INTENTS",
    "PRICE",
    "LOGISTICS",
    "QUALITY",
    "SPEC",
    "ESCALATION",
    "price",
    "logistics",
    "quality",
    "spec",
    "escalation",
)


def main() -> int:
    """Inspect P0 classifier behavior."""

    print("=" * 80)
    print("inspecting Phase 3-I-I intent classifier P0 behavior")

    errors: list[str] = []

    workbook_cases = load_p0_workbook_cases(errors=errors)
    behavior_result = inspect_classifier_behavior(workbook_cases)
    source_result = inspect_source_windows(errors=errors)

    result = {
        "case_count": len(workbook_cases),
        "behavior_result": behavior_result,
        "source_result": source_result,
        "recommended_patch_rules": build_recommended_patch_rules(behavior_result),
        "errors": errors,
    }

    pprint(result)

    if errors:
        print("Phase 3-I-I intent classifier P0 behavior inspection failed")
        return 1

    print("Phase 3-I-I intent classifier P0 behavior inspection passed")
    return 0


def load_p0_workbook_cases(
    *,
    errors: list[str],
) -> list[dict[str, Any]]:
    """Load P0 cases from workbook."""

    if not TEST_CASES_FILE.exists():
        errors.append(f"missing workbook: {TEST_CASES_FILE}")
        return []

    workbook = load_workbook(TEST_CASES_FILE, data_only=True)
    sheet = workbook["test_cases"]

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]

    cases: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        case_id = str(row_data.get("case_id") or "")

        if case_id in P0_CASE_IDS:
            cases.append(row_data)

    if len(cases) != len(P0_CASE_IDS):
        errors.append(f"P0 case count mismatch: {len(cases)}")

    return cases


def inspect_classifier_behavior(
    cases: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Inspect keyword classifier behavior for each P0 case."""

    rows: list[dict[str, Any]] = []

    for case in cases:
        query = str(case.get("input_message") or "")
        expected_intent = str(case.get("expected_intent") or case.get("category") or "")
        keyword_result = classify_intent_by_keywords(query)

        rows.append(
            {
                "case_id": case.get("case_id"),
                "query": query,
                "expected_intent": expected_intent,
                "keyword_intent": keyword_result.intent,
                "keyword_confidence": keyword_result.confidence,
                "keyword_reason": keyword_result.reason,
                "keyword_metadata": keyword_result.metadata,
                "signals": detect_patch_signals(query),
                "expected_handoff": case.get("expected_handoff"),
                "scenario_type": case.get("scenario_type"),
                "must_contain_all": case.get("must_contain_all"),
            }
        )

    return rows


def inspect_source_windows(
    *,
    errors: list[str],
) -> dict[str, Any]:
    """Inspect intent classifier source windows."""

    if not INTENT_CLASSIFIER_FILE.exists():
        errors.append(f"missing classifier file: {INTENT_CLASSIFIER_FILE}")
        return {"exists": False}

    content = INTENT_CLASSIFIER_FILE.read_text(encoding="utf-8")
    lines = content.splitlines()

    windows: dict[str, list[str]] = {}

    for anchor in SOURCE_ANCHORS:
        line_number = find_line_number(lines, anchor)

        if line_number is None:
            continue

        windows[anchor] = source_window(
            lines=lines,
            center_line=line_number,
            before=12,
            after=70,
        )

    return {
        "exists": True,
        "path": str(INTENT_CLASSIFIER_FILE.relative_to(BACKEND_ROOT)),
        "line_count": len(lines),
        "windows": windows,
        "string_presence": {
            "escalation": "escalation" in content,
            "price": "price" in content,
            "logistics": "logistics" in content,
            "quality": "quality" in content,
            "spec": "spec" in content,
            "老客户": "老客户" in content,
            "实在价": "实在价" in content,
            "质检报告": "质检报告" in content,
            "定制": "定制" in content,
            "LOGO": "LOGO" in content,
            "全部参数": "全部参数" in content,
            "最长": "最长" in content,
        },
    }


def detect_patch_signals(query: str) -> list[str]:
    """Detect expected patch signals."""

    signal_rules = {
        "escalation_high_priority": (
            "投诉",
            "差评",
            "骗子",
            "赔不赔",
            "赔付",
            "赔",
            "定制",
            "LOGO",
        ),
        "price_high_priority": (
            "报价",
            "报个价",
            "批发价",
            "批发",
            "实在价",
            "老客户",
            "多少钱",
            "价格",
            "100个",
            "1000个",
        ),
        "logistics_high_priority": (
            "顺丰",
            "新疆",
            "澳门",
            "运费",
            "差价",
            "补多少钱",
            "发货",
        ),
        "quality_high_priority": (
            "原厂",
            "OEM正品",
            "质检",
            "认证",
            "报告",
            "哪个更好",
            "区别",
            "会不会",
            "掉色",
            "发霉",
            "褪色",
            "夜光",
            "蓄光",
        ),
        "spec_high_priority": (
            "全部参数",
            "哪些款",
            "哪款",
            "最长",
            "杆",
            "螺纹",
            "球径",
            "锥度",
        ),
    }

    matched: list[str] = []

    for signal, fragments in signal_rules.items():
        if any(fragment in query for fragment in fragments):
            matched.append(signal)

    return matched


def build_recommended_patch_rules(
    behavior_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build recommended patch rules from behavior rows."""

    return [
        {
            "priority": 1,
            "target_intent": "escalation",
            "rule": "投诉/差评/骗子/赔付/定制/LOGO should override all other intents",
            "affected_cases": [
                row["case_id"]
                for row in behavior_rows
                if row["expected_intent"] == "escalation"
            ],
        },
        {
            "priority": 2,
            "target_intent": "price",
            "rule": "报价/批发价/实在价/老客户/多少钱 should override SKU/spec lookup",
            "affected_cases": [
                row["case_id"]
                for row in behavior_rows
                if row["expected_intent"] == "price"
            ],
        },
        {
            "priority": 3,
            "target_intent": "logistics",
            "rule": "顺丰/新疆/运费/差价/补多少钱 should override price fallback",
            "affected_cases": [
                row["case_id"]
                for row in behavior_rows
                if row["expected_intent"] == "logistics"
            ],
        },
        {
            "priority": 4,
            "target_intent": "quality",
            "rule": "原厂/质检/认证/会不会/哪个更好 should route quality",
            "affected_cases": [
                row["case_id"]
                for row in behavior_rows
                if row["expected_intent"] == "quality"
            ],
        },
        {
            "priority": 5,
            "target_intent": "spec",
            "rule": "全部参数/哪些款/哪款/最长/杆 should route spec even with material words",
            "affected_cases": [
                row["case_id"]
                for row in behavior_rows
                if row["expected_intent"] == "spec"
            ],
        },
    ]


def find_line_number(lines: list[str], anchor: str) -> int | None:
    """Find 1-based line number."""

    for index, line in enumerate(lines, start=1):
        if anchor in line:
            return index

    return None


def source_window(
    *,
    lines: list[str],
    center_line: int,
    before: int,
    after: int,
) -> list[str]:
    """Return source window."""

    start = max(center_line - before, 1)
    end = min(center_line + after, len(lines))

    return [
        f"{line_number}: {lines[line_number - 1]}"
        for line_number in range(start, end + 1)
    ]


if __name__ == "__main__":
    raise SystemExit(main())