# ruff: noqa: E402,I001
"""Check Phase 3-I-B quality_questions.xlsx file.

This script verifies the fixed Quality QA input file before building real
quality knowledge chunks.
"""

from __future__ import annotations

import re
from pathlib import Path
from pprint import pprint
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

QUALITY_FILE: Final[Path] = (
    PROJECT_ROOT
    / "data"
    / "uploads"
    / "qa_pairs"
    / "quality_questions.xlsx"
)

SKU_MASTER_FILE: Final[Path] = (
    PROJECT_ROOT
    / "data"
    / "uploads"
    / "specs"
    / "sku_master.xlsx"
)

QA_SHEET_NAME: Final[str] = "qa_pairs"
LABEL_SHEET_NAME: Final[str] = "label_dictionary"

REQUIRED_QA_COLUMNS: Final[tuple[str, ...]] = (
    "qa_id",
    "source_group_id",
    "question_raw",
    "question_normalized",
    "answer_raw",
    "answer_standard",
    "primary_intent",
    "intent_subtype",
    "secondary_intents",
    "related_sku_ids",
    "risk_flags",
    "handoff_required",
    "verification_status",
    "expected_source",
    "notes",
    "metadata",
)

ALLOWED_PRIMARY_INTENTS: Final[set[str]] = {"quality"}
ALLOWED_HANDOFF_VALUES: Final[set[str]] = {"true", "false"}
ALLOWED_VERIFICATION_STATUS: Final[set[str]] = {
    "pending",
    "verified",
    "rejected",
}
ALLOWED_RISK_FLAGS: Final[set[str]] = {
    "",
    "unsupported_claim",
    "safety_claim",
    "quality_guarantee",
    "durability_claim",
    "certification_claim",
    "inspection_report_required",
    "aftersale_commitment",
    "compensation_claim",
    "installation_advice",
    "vehicle_fitment_unverified",
    "data_conflict",
    "business_policy_confirmation",
}


def load_sheet(
    *,
    workbook_path: Path,
    sheet_name: str,
) -> Worksheet:
    """Load one workbook sheet."""

    workbook = load_workbook(workbook_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"required sheet not found: {sheet_name}; "
            f"available={workbook.sheetnames}"
        )

    return workbook[sheet_name]


def read_rows(
    sheet: Worksheet,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Read worksheet into headers and row dicts."""

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise RuntimeError(f"sheet is empty: {sheet.title}")

    headers = [
        str(value).strip()
        for value in rows[0]
        if value is not None and str(value).strip()
    ]

    if not headers:
        raise RuntimeError(f"sheet header is empty: {sheet.title}")

    data_rows: list[dict[str, Any]] = []

    for row_index, values in enumerate(rows[1:], start=2):
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
        }

        if all(_is_blank(value) for value in row.values()):
            continue

        row["_excel_row"] = row_index
        data_rows.append(row)

    return headers, data_rows


def load_sku_ids() -> set[str]:
    """Load known SKU IDs from sku_master.xlsx when available."""

    if not SKU_MASTER_FILE.exists():
        return set()

    sheet = load_workbook(SKU_MASTER_FILE, data_only=True).active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return set()

    headers = [
        str(value).strip()
        for value in rows[0]
        if value is not None and str(value).strip()
    ]

    if "SKU_ID" not in headers:
        return set()

    sku_index = headers.index("SKU_ID")
    result: set[str] = set()

    for values in rows[1:]:
        if sku_index >= len(values):
            continue

        value = values[sku_index]

        if value is None:
            continue

        sku_id = str(value).strip()

        if sku_id:
            result.add(sku_id)

    return result


def check_quality_file() -> bool:
    """Check quality_questions.xlsx."""

    print("=" * 80)
    print("checking quality_questions.xlsx")
    print(f"quality_file={QUALITY_FILE}")
    print(f"sku_master_file={SKU_MASTER_FILE}")

    errors: list[str] = []
    warnings: list[str] = []

    if not QUALITY_FILE.exists():
        print("failed: quality_questions.xlsx not found")
        return False

    qa_sheet = load_sheet(
        workbook_path=QUALITY_FILE,
        sheet_name=QA_SHEET_NAME,
    )
    workbook = qa_sheet.parent

    if LABEL_SHEET_NAME not in workbook.sheetnames:
        errors.append("label_dictionary sheet missing")

    headers, rows = read_rows(qa_sheet)

    missing_columns = [
        column
        for column in REQUIRED_QA_COLUMNS
        if column not in headers
    ]

    extra_columns = [
        column
        for column in headers
        if column not in REQUIRED_QA_COLUMNS
    ]

    if missing_columns:
        errors.append(f"missing required columns: {missing_columns}")

    if extra_columns:
        warnings.append(f"extra columns detected: {extra_columns}")

    if not rows:
        errors.append("qa_pairs has no data rows")

    sku_ids = load_sku_ids()

    if not sku_ids:
        warnings.append(
            "sku_master.xlsx missing or unreadable; related_sku_ids existence "
            "check skipped"
        )

    qa_ids_seen: set[str] = set()
    source_group_ids_seen: set[str] = set()
    related_sku_values: set[str] = set()
    intent_subtypes: set[str] = set()
    risk_flags_seen: set[str] = set()
    handoff_values_seen: set[str] = set()
    verification_status_seen: set[str] = set()

    for row in rows:
        excel_row = int(row.get("_excel_row") or 0)

        qa_id = _text(row.get("qa_id"))
        source_group_id = _text(row.get("source_group_id"))
        question_raw = _text(row.get("question_raw"))
        question_normalized = _text(row.get("question_normalized"))
        answer_raw = _text(row.get("answer_raw"))
        answer_standard = _text(row.get("answer_standard"))
        primary_intent = _text(row.get("primary_intent"))
        intent_subtype = _text(row.get("intent_subtype"))
        related_sku_ids = _split_semicolon(row.get("related_sku_ids"))
        risk_flags = _split_semicolon(row.get("risk_flags"))
        handoff_required = _text(row.get("handoff_required")).lower()
        verification_status = _text(row.get("verification_status")).lower()

        if not re.fullmatch(r"QUAL\d{4}", qa_id):
            errors.append(f"row {excel_row}: invalid qa_id={qa_id!r}")

        if qa_id in qa_ids_seen:
            errors.append(f"row {excel_row}: duplicated qa_id={qa_id!r}")
        qa_ids_seen.add(qa_id)

        if not source_group_id.startswith("QUAL_SRC"):
            errors.append(
                f"row {excel_row}: source_group_id must start with QUAL_SRC"
            )
        source_group_ids_seen.add(source_group_id)

        if not question_raw:
            errors.append(f"row {excel_row}: question_raw is empty")

        if not question_normalized:
            errors.append(f"row {excel_row}: question_normalized is empty")

        if not answer_raw:
            errors.append(f"row {excel_row}: answer_raw is empty")

        if not answer_standard:
            errors.append(f"row {excel_row}: answer_standard is empty")

        if primary_intent not in ALLOWED_PRIMARY_INTENTS:
            errors.append(
                f"row {excel_row}: primary_intent must be quality, "
                f"got {primary_intent!r}"
            )

        if not intent_subtype:
            errors.append(f"row {excel_row}: intent_subtype is empty")
        else:
            intent_subtypes.add(intent_subtype)

        for sku_id in related_sku_ids:
            related_sku_values.add(sku_id)

            if sku_ids and sku_id not in sku_ids:
                errors.append(
                    f"row {excel_row}: related_sku_id not found in "
                    f"sku_master.xlsx: {sku_id}"
                )

        for risk_flag in risk_flags:
            risk_flags_seen.add(risk_flag)

            if risk_flag not in ALLOWED_RISK_FLAGS:
                errors.append(
                    f"row {excel_row}: unsupported risk_flag={risk_flag!r}"
                )

        handoff_values_seen.add(handoff_required)

        if handoff_required not in ALLOWED_HANDOFF_VALUES:
            errors.append(
                f"row {excel_row}: handoff_required must be true/false, "
                f"got {handoff_required!r}"
            )

        verification_status_seen.add(verification_status)

        if verification_status not in ALLOWED_VERIFICATION_STATUS:
            errors.append(
                f"row {excel_row}: invalid verification_status="
                f"{verification_status!r}"
            )

        if (
            "guarantee" in " ".join(risk_flags)
            or "claim" in " ".join(risk_flags)
            or "safety_claim" in risk_flags
        ) and handoff_required != "true":
            warnings.append(
                f"row {excel_row}: risk_flags={risk_flags} may require "
                "handoff_required=true"
            )

    qa_number_errors = _check_qa_id_continuity(qa_ids_seen)

    errors.extend(qa_number_errors)

    summary = {
        "quality_file_exists": QUALITY_FILE.exists(),
        "workbook_sheets": workbook.sheetnames,
        "qa_sheet": QA_SHEET_NAME,
        "qa_row_count": len(rows),
        "qa_id_count": len(qa_ids_seen),
        "source_group_id_count": len(source_group_ids_seen),
        "intent_subtypes": sorted(intent_subtypes),
        "risk_flags_seen": sorted(risk_flags_seen),
        "handoff_values_seen": sorted(handoff_values_seen),
        "verification_status_seen": sorted(verification_status_seen),
        "related_sku_count": len(related_sku_values),
        "sku_master_exists": SKU_MASTER_FILE.exists(),
        "sku_master_sku_count": len(sku_ids),
        "warnings": warnings,
        "errors": errors,
    }

    pprint(summary)

    if errors:
        print("quality_questions.xlsx check failed")
        return False

    print("quality_questions.xlsx check passed")
    return True


def _check_qa_id_continuity(
    qa_ids: set[str],
) -> list[str]:
    """Check QUAL ID continuity."""

    if not qa_ids:
        return ["qa_id set is empty"]

    numbers: list[int] = []

    for qa_id in qa_ids:
        match = re.fullmatch(r"QUAL(\d{4})", qa_id)

        if match:
            numbers.append(int(match.group(1)))

    if not numbers:
        return ["no valid QUAL qa_id found"]

    expected = set(range(min(numbers), max(numbers) + 1))
    actual = set(numbers)
    missing = sorted(expected - actual)

    if missing:
        return [f"qa_id sequence missing numbers: {missing}"]

    return []


def _split_semicolon(
    value: object,
) -> list[str]:
    """Split semicolon separated text."""

    text = _text(value)

    if not text:
        return []

    return [
        item.strip()
        for item in text.split(";")
        if item.strip()
    ]


def _text(
    value: object,
) -> str:
    """Return stripped text."""

    if value is None:
        return ""

    return str(value).strip()


def _is_blank(
    value: object,
) -> bool:
    """Return whether value is blank."""

    return value is None or str(value).strip() == ""


def main() -> int:
    """Run quality file check."""

    try:
        passed = check_quality_file()
    except Exception as exc:
        print(f"quality_questions.xlsx check crashed: {type(exc).__name__}: {exc}")
        return 1

    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())