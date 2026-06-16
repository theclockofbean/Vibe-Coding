"""Check Phase 3-I-I evaluator risk case semantics."""

from __future__ import annotations

import importlib.util
from pathlib import Path
from pprint import pprint
from typing import Any, Final, Protocol

from openpyxl import load_workbook


BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

EVAL_SCRIPT: Final[Path] = BACKEND_ROOT / "scripts/check_phase3ii_real_llm_50_case_eval.py"
TEST_CASES_FILE: Final[Path] = PROJECT_ROOT / "data/evaluation/test_cases_draft.xlsx"

TARGET_CASE_IDS: Final[set[str]] = {
    "TC_SPEC_010",
    "TC_SPEC_015",
    "TC_SPEC_017",
}


class EvalModuleProtocol(Protocol):
    """Protocol for imported eval module."""

    def is_risk_case(
        self,
        *,
        scenario_type: str,
        expected_handoff: bool,
        is_critical: bool,
    ) -> bool:
        """Return whether one case should trigger risk gate."""


def main() -> int:
    """Run evaluator risk case semantic checks."""

    print("=" * 80)
    print("checking Phase 3-I-I evaluator risk case semantics")

    errors: list[str] = []

    source_result = inspect_source(errors=errors)
    behavior_result = inspect_behavior(errors=errors)
    workbook_result = inspect_workbook_targets(errors=errors)

    result = {
        "source_result": source_result,
        "behavior_result": behavior_result,
        "workbook_result": workbook_result,
        "errors": errors,
    }

    pprint(result)

    if errors:
        print("Phase 3-I-I evaluator risk case semantics check failed")
        return 1

    print("Phase 3-I-I evaluator risk case semantics check passed")
    return 0


def inspect_source(
    *,
    errors: list[str],
) -> dict[str, Any]:
    """Inspect patched evaluator source."""

    if not EVAL_SCRIPT.exists():
        errors.append(f"missing eval script: {EVAL_SCRIPT}")
        return {"exists": False}

    content = EVAL_SCRIPT.read_text(encoding="utf-8")

    old_semantics_present = "return scenario_type == \"risk\" or is_critical" in content
    new_semantics_present = (
        "return scenario_type == \"risk\" or expected_handoff" in content
    )
    call_has_expected_handoff = "expected_handoff=expected_handoff" in content

    if old_semantics_present:
        errors.append("old is_critical risk semantics still present")

    if not new_semantics_present:
        errors.append("new expected_handoff risk semantics not found")

    if not call_has_expected_handoff:
        errors.append("is_risk_case call does not pass expected_handoff")

    return {
        "old_semantics_present": old_semantics_present,
        "new_semantics_present": new_semantics_present,
        "call_has_expected_handoff": call_has_expected_handoff,
    }


def inspect_behavior(
    *,
    errors: list[str],
) -> dict[str, Any]:
    """Import evaluator and check behavior."""

    module = load_eval_module()

    checks = {
        "core_critical_no_handoff": module.is_risk_case(
            scenario_type="core",
            expected_handoff=False,
            is_critical=True,
        ),
        "risk_no_handoff": module.is_risk_case(
            scenario_type="risk",
            expected_handoff=False,
            is_critical=False,
        ),
        "core_expected_handoff": module.is_risk_case(
            scenario_type="core",
            expected_handoff=True,
            is_critical=False,
        ),
        "boundary_critical_no_handoff": module.is_risk_case(
            scenario_type="boundary",
            expected_handoff=False,
            is_critical=True,
        ),
    }

    if checks["core_critical_no_handoff"] is not False:
        errors.append("core critical non-handoff case should not be risk-gated")

    if checks["boundary_critical_no_handoff"] is not False:
        errors.append("boundary critical non-handoff case should not be risk-gated")

    if checks["risk_no_handoff"] is not True:
        errors.append("risk scenario should be risk-gated")

    if checks["core_expected_handoff"] is not True:
        errors.append("expected_handoff case should be risk-gated")

    return checks


def inspect_workbook_targets(
    *,
    errors: list[str],
) -> list[dict[str, Any]]:
    """Verify target cases are critical but not expected handoff."""

    workbook = load_workbook(TEST_CASES_FILE, data_only=True)
    sheet = workbook["test_cases"]

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]

    rows: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }

        case_id = str(row_data.get("case_id") or "")

        if case_id not in TARGET_CASE_IDS:
            continue

        compact = {
            "case_id": case_id,
            "scenario_type": row_data.get("scenario_type"),
            "expected_handoff": row_data.get("expected_handoff"),
            "is_critical": row_data.get("is_critical"),
            "must_contain_all": row_data.get("must_contain_all"),
        }
        rows.append(compact)

        if str(row_data.get("expected_handoff")).upper() != "FALSE":
            errors.append(f"{case_id}: expected_handoff should be FALSE")

        if str(row_data.get("is_critical")).upper() != "TRUE":
            errors.append(f"{case_id}: is_critical should be TRUE")

    if len(rows) != len(TARGET_CASE_IDS):
        errors.append(f"target rows count mismatch: {len(rows)}")

    return rows


def load_eval_module() -> EvalModuleProtocol:
    """Load evaluator script as module."""

    spec = importlib.util.spec_from_file_location(
        "phase3ii_eval_module",
        EVAL_SCRIPT,
    )

    if spec is None or spec.loader is None:
        raise RuntimeError("cannot load eval script spec")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    return module  # type: ignore[return-value]


if __name__ == "__main__":
    raise SystemExit(main())