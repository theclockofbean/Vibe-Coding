# ruff: noqa: E402,I001
"""Check spec_questions.xlsx source file for Spec KB."""

from __future__ import annotations

from pathlib import Path
from pprint import pprint
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]

BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

SPEC_FILE_CANDIDATES: Final[tuple[Path, ...]] = (
    PROJECT_ROOT / "data/uploads/conversations/qa_pairs_raw/spec_questions.xlsx",
    PROJECT_ROOT / "data/uploads/qa_pairs/spec_questions.xlsx",
    PROJECT_ROOT / "data/uploads/spec_questions.xlsx",
)

EXPECTED_SHEET_NAME: Final[str] = "qa_pairs"
EXPECTED_ROW_COUNT: Final[int] = 23
EXPECTED_PRIMARY_INTENT: Final[str] = "spec"

REQUIRED_COLUMNS: Final[set[str]] = {
    "qa_id",
    "source_group_id",
    "primary_intent",
    "secondary_intents",
    "intent_subtype",
    "question_raw",
    "question_normalized",
    "answer_raw",
    "answer_standard",
    "related_sku_ids",
    "required_fields",
    "answer_source",
    "handoff_required",
    "risk_flags",
    "verification_status",
    "review_notes",
}


def main() -> int:
    """Run Spec source file check."""

    print("=" * 80)
    print("checking spec_questions.xlsx source file")

    spec_file = find_spec_file()
    workbook = load_workbook(spec_file, read_only=True, data_only=True)

    errors: list[str] = []

    if EXPECTED_SHEET_NAME not in workbook.sheetnames:
        errors.append(f"missing sheet: {EXPECTED_SHEET_NAME}")
        pprint(
            {
                "spec_file": str(spec_file),
                "sheetnames": workbook.sheetnames,
                "errors": errors,
            }
        )
        return 1

    worksheet = workbook[EXPECTED_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        errors.append("qa_pairs sheet is empty")
        pprint({"spec_file": str(spec_file), "errors": errors})
        return 1

    headers = [
        str(value).strip() if value is not None else ""
        for value in rows[0]
    ]

    records = [
        dict(zip(headers, row, strict=False))
        for row in rows[1:]
        if any(value is not None and str(value).strip() for value in row)
    ]

    qa_ids = [
        str(record.get("qa_id", "")).strip()
        for record in records
    ]
    primary_intents = {
        str(record.get("primary_intent", "")).strip()
        for record in records
    }
    intent_subtypes = sorted(
        {
            str(record.get("intent_subtype", "")).strip()
            for record in records
            if str(record.get("intent_subtype", "")).strip()
        }
    )
    risk_flags = sorted(
        {
            str(record.get("risk_flags", "")).strip()
            for record in records
            if str(record.get("risk_flags", "")).strip()
        }
    )
    handoff_values = sorted(
        {
            str(record.get("handoff_required", "")).strip()
            for record in records
            if str(record.get("handoff_required", "")).strip()
        }
    )

    validate_basic_shape(
        headers=headers,
        records=records,
        qa_ids=qa_ids,
        primary_intents=primary_intents,
        errors=errors,
    )
    validate_content(records=records, errors=errors)

    result: dict[str, Any] = {
        "spec_file": str(spec_file),
        "sheet_name": EXPECTED_SHEET_NAME,
        "record_count": len(records),
        "headers": headers,
        "qa_id_first": qa_ids[:3],
        "qa_id_last": qa_ids[-3:],
        "primary_intents": sorted(primary_intents),
        "intent_subtypes": intent_subtypes,
        "risk_flags": risk_flags,
        "handoff_required_values": handoff_values,
        "errors": errors,
    }

    pprint(result)

    if errors:
        print("spec_questions.xlsx source file check failed")
        return 1

    print("spec_questions.xlsx source file check passed")
    return 0


def validate_basic_shape(
    *,
    headers: list[str],
    records: list[dict[str, Any]],
    qa_ids: list[str],
    primary_intents: set[str],
    errors: list[str],
) -> None:
    """Validate file shape."""

    if len(records) != EXPECTED_ROW_COUNT:
        errors.append(
            f"expected {EXPECTED_ROW_COUNT} records, got {len(records)}"
        )

    missing_columns = sorted(REQUIRED_COLUMNS - set(headers))

    if missing_columns:
        errors.append(f"missing columns: {missing_columns}")

    if len(qa_ids) != len(set(qa_ids)):
        errors.append("duplicated qa_id found")

    if "" in qa_ids:
        errors.append("empty qa_id found")

    expected_qa_ids = [
        f"SPEC{index:04d}"
        for index in range(1, EXPECTED_ROW_COUNT + 1)
    ]

    if qa_ids != expected_qa_ids:
        errors.append("qa_id sequence must be SPEC0001-SPEC0023")

    if primary_intents != {EXPECTED_PRIMARY_INTENT}:
        errors.append(
            f"primary_intent must only be {EXPECTED_PRIMARY_INTENT}, "
            f"got {sorted(primary_intents)}"
        )


def validate_content(
    *,
    records: list[dict[str, Any]],
    errors: list[str],
) -> None:
    """Validate key content fields."""

    required_non_empty_columns = (
        "qa_id",
        "primary_intent",
        "question_raw",
        "question_normalized",
        "answer_raw",
        "answer_standard",
        "answer_source",
        "handoff_required",
        "risk_flags",
        "verification_status",
    )

    for row_index, record in enumerate(records, start=2):
        for column in required_non_empty_columns:
            value = str(record.get(column, "")).strip()
            if not value:
                errors.append(f"row {row_index}: empty {column}")


def find_spec_file() -> Path:
    """Find spec_questions.xlsx."""

    for candidate in SPEC_FILE_CANDIDATES:
        if candidate.exists():
            return candidate

    searched = "\n".join(str(path) for path in SPEC_FILE_CANDIDATES)
    raise FileNotFoundError(
        f"spec_questions.xlsx not found. searched:\n{searched}"
    )


if __name__ == "__main__":
    raise SystemExit(main())