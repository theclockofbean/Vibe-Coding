"""Inspect P0 module routing failures for Phase 3-I-I major gate."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from pprint import pprint
from typing import Any, Final

from openpyxl import load_workbook


BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

REPORT_FILE: Final[Path] = (
    PROJECT_ROOT / "logs/evaluation/phase3ii_real_llm_50_case_eval_report.json"
)
TEST_CASES_FILE: Final[Path] = PROJECT_ROOT / "data/evaluation/test_cases_draft.xlsx"
APP_AGENT_DIR: Final[Path] = BACKEND_ROOT / "app/agent"

P0_CASE_IDS: Final[tuple[str, ...]] = (
    "TC_SPEC_004",
    "TC_SPEC_007",
    "TC_SPEC_011",
    "TC_LOGI_002",
    "TC_LOGI_003",
    "TC_PRICE_002",
    "TC_PRICE_007",
    "TC_PRICE_008",
    "TC_QUAL_001",
    "TC_QUAL_003",
    "TC_QUAL_004",
    "TC_QUAL_005",
    "TC_QUAL_008",
    "TC_ESCA_001",
    "TC_ESCA_002",
    "TC_ESCA_003",
)

ROUTING_MARKERS: Final[tuple[str, ...]] = (
    "selected_module",
    "candidate_modules",
    "expected_module",
    "route",
    "routing",
    "intent",
    "classifier",
    "conflict_type",
    "spec",
    "price",
    "quality",
    "logistics",
    "escalation",
)

QUERY_SIGNAL_GROUPS: Final[dict[str, tuple[str, ...]]] = {
    "sku_exact": ("SKU",),
    "spec_dimension": (
        "螺纹",
        "杆长",
        "球径",
        "锥度",
        "尺寸",
        "M8",
        "M10",
        "M12",
        "M14",
        "mm",
    ),
    "material_quality": (
        "材质",
        "铝合金",
        "不锈钢",
        "碳纤维",
        "真皮",
        "钛合金",
        "原厂",
        "质检",
        "认证",
        "报告",
        "夜光",
        "发霉",
        "掉色",
    ),
    "price": (
        "价格",
        "报价",
        "批发",
        "优惠",
        "便宜",
        "实在价",
        "老客户",
        "1000",
        "500",
    ),
    "logistics": (
        "发货",
        "顺丰",
        "新疆",
        "澳门",
        "运费",
        "差价",
        "破损",
        "退换货",
        "换货",
    ),
    "escalation": (
        "投诉",
        "差评",
        "骗子",
        "客服",
        "定制",
        "LOGO",
        "赔付",
        "人工",
    ),
}


def main() -> int:
    """Inspect routing failures."""

    print("=" * 80)
    print("inspecting Phase 3-I-I P0 module routing failures")

    errors: list[str] = []

    report_cases = load_report_cases(errors=errors)
    workbook_cases = load_workbook_cases(errors=errors)

    result = {
        "p0_case_count": len(report_cases),
        "by_expected_selected": build_expected_selected_counts(report_cases),
        "by_signal_group": build_signal_group_counts(workbook_cases),
        "case_matrix": build_case_matrix(report_cases, workbook_cases),
        "routing_source_candidates": inspect_routing_source_candidates(),
        "recommended_fix_hypotheses": build_fix_hypotheses(report_cases, workbook_cases),
        "errors": errors,
    }

    pprint(result)

    if errors:
        print("Phase 3-I-I P0 module routing failure inspection failed")
        return 1

    print("Phase 3-I-I P0 module routing failure inspection passed")
    return 0


def load_report_cases(
    *,
    errors: list[str],
) -> list[dict[str, Any]]:
    """Load P0 cases from latest report."""

    if not REPORT_FILE.exists():
        errors.append(f"missing report file: {REPORT_FILE}")
        return []

    data = json.loads(REPORT_FILE.read_text(encoding="utf-8"))
    results = data.get("results", [])

    if not isinstance(results, list):
        errors.append("report.results must be list")
        return []

    selected: list[dict[str, Any]] = []

    for item in results:
        if not isinstance(item, dict):
            continue

        case_id = str(item.get("case_id") or "")

        if case_id not in P0_CASE_IDS:
            continue

        selected.append(
            {
                "case_id": case_id,
                "category": item.get("category"),
                "expected_module": item.get("expected_module"),
                "selected_module": item.get("selected_module"),
                "scenario_type": item.get("scenario_type"),
                "answer_strategy_mode": item.get("answer_strategy_mode"),
                "failure_reasons": item.get("failure_reasons", []),
            }
        )

    return selected


def load_workbook_cases(
    *,
    errors: list[str],
) -> dict[str, dict[str, Any]]:
    """Load P0 cases from workbook."""

    if not TEST_CASES_FILE.exists():
        errors.append(f"missing workbook file: {TEST_CASES_FILE}")
        return {}

    workbook = load_workbook(TEST_CASES_FILE, data_only=True)
    sheet = workbook["test_cases"]

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]

    rows: dict[str, dict[str, Any]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        case_id = str(row_data.get("case_id") or "")

        if case_id in P0_CASE_IDS:
            rows[case_id] = row_data

    missing = sorted(set(P0_CASE_IDS) - set(rows))

    if missing:
        errors.append(f"missing P0 cases from workbook: {missing}")

    return rows


def build_expected_selected_counts(
    report_cases: list[dict[str, Any]],
) -> dict[str, int]:
    """Build expected -> selected counts."""

    counts: Counter[str] = Counter()

    for item in report_cases:
        expected = str(item.get("expected_module"))
        selected = str(item.get("selected_module"))
        counts[f"{expected} -> {selected}"] += 1

    return dict(counts)


def build_signal_group_counts(
    workbook_cases: dict[str, dict[str, Any]],
) -> dict[str, int]:
    """Count signal groups in P0 queries."""

    counts: Counter[str] = Counter()

    for row in workbook_cases.values():
        query = str(row.get("input_message") or "")

        for group, fragments in QUERY_SIGNAL_GROUPS.items():
            if any(fragment in query for fragment in fragments):
                counts[group] += 1

    return dict(counts)


def build_case_matrix(
    report_cases: list[dict[str, Any]],
    workbook_cases: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build compact matrix for each P0 case."""

    rows: list[dict[str, Any]] = []

    for item in report_cases:
        case_id = str(item["case_id"])
        workbook_row = workbook_cases.get(case_id, {})
        query = str(workbook_row.get("input_message") or "")

        rows.append(
            {
                "case_id": case_id,
                "query": query,
                "category": item.get("category"),
                "expected_module": item.get("expected_module"),
                "selected_module": item.get("selected_module"),
                "expected_handoff": workbook_row.get("expected_handoff"),
                "scenario_type": workbook_row.get("scenario_type"),
                "is_critical": workbook_row.get("is_critical"),
                "signals": detect_query_signals(query),
                "must_contain_all": workbook_row.get("must_contain_all"),
                "failure_reasons": item.get("failure_reasons", [])[:5],
            }
        )

    return rows


def detect_query_signals(
    query: str,
) -> list[str]:
    """Detect simple query signal groups."""

    matched: list[str] = []

    for group, fragments in QUERY_SIGNAL_GROUPS.items():
        if any(fragment in query for fragment in fragments):
            matched.append(group)

    return matched


def inspect_routing_source_candidates() -> list[dict[str, Any]]:
    """Find likely routing source files and related lines."""

    if not APP_AGENT_DIR.exists():
        return []

    candidates: list[dict[str, Any]] = []

    for path in sorted(APP_AGENT_DIR.rglob("*.py")):
        if "__pycache__" in path.parts:
            continue

        content = path.read_text(encoding="utf-8")
        lowered = content.lower()

        if not any(marker.lower() in lowered for marker in ROUTING_MARKERS):
            continue

        lines = extract_lines(
            content=content,
            markers=ROUTING_MARKERS,
            limit=80,
        )

        candidates.append(
            {
                "path": str(path.relative_to(BACKEND_ROOT)),
                "line_count": len(content.splitlines()),
                "matched_line_count": len(lines),
                "lines": lines[:40],
            }
        )

    return candidates[:20]


def extract_lines(
    *,
    content: str,
    markers: tuple[str, ...],
    limit: int,
) -> list[str]:
    """Extract lines containing routing markers."""

    lines: list[str] = []

    for line_number, line in enumerate(content.splitlines(), start=1):
        lowered = line.lower()

        if any(marker.lower() in lowered for marker in markers):
            lines.append(f"{line_number}: {line.rstrip()}")

    return lines[:limit]


def build_fix_hypotheses(
    report_cases: list[dict[str, Any]],
    workbook_cases: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build preliminary fix hypotheses."""

    matrix = build_case_matrix(report_cases, workbook_cases)

    hypotheses: list[dict[str, Any]] = []

    hypotheses.append(
        {
            "hypothesis": "SKU + spec dimension signals should override quality/material explanations",
            "affected_cases": [
                item["case_id"]
                for item in matrix
                if item["expected_module"] == "spec"
                and item["selected_module"] in ("quality", None)
            ],
        }
    )
    hypotheses.append(
        {
            "hypothesis": "price quote and bulk quantity signals should override spec SKU lookup",
            "affected_cases": [
                item["case_id"]
                for item in matrix
                if item["expected_module"] == "price"
                and item["selected_module"] in ("spec", None)
            ],
        }
    )
    hypotheses.append(
        {
            "hypothesis": "logistics fee/region signals should override price/spec fallback",
            "affected_cases": [
                item["case_id"]
                for item in matrix
                if item["expected_module"] == "logistics"
                and item["selected_module"] in ("price", None)
            ],
        }
    )
    hypotheses.append(
        {
            "hypothesis": "quality certification/material comparison signals need explicit quality route",
            "affected_cases": [
                item["case_id"]
                for item in matrix
                if item["expected_module"] == "quality"
                and item["selected_module"] in ("spec", None)
            ],
        }
    )
    hypotheses.append(
        {
            "hypothesis": "complaint/customization/compensation should map to escalation intent",
            "affected_cases": [
                item["case_id"]
                for item in matrix
                if item["expected_module"] == "escalation"
                and item["selected_module"] != "escalation"
            ],
        }
    )

    return hypotheses


if __name__ == "__main__":
    raise SystemExit(main())