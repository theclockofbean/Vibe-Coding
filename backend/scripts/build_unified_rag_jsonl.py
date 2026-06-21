from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook


QA_SOURCES = [
    {
        "domain": "spec",
        "source_type": "qa",
        "input": Path("../data/uploads/conversations/qa_pairs_raw/spec_questions.xlsx"),
        "output": Path("../data/processed/rag_unified/spec_qa.jsonl"),
    },
    {
        "domain": "logistics",
        "source_type": "qa",
        "input": Path("../data/uploads/conversations/qa_pairs_raw/logistics_questions.xlsx"),
        "output": Path("../data/processed/rag_unified/logistics_qa.jsonl"),
    },
    {
        "domain": "price",
        "source_type": "qa",
        "input": Path("../data/uploads/conversations/qa_pairs_raw/price_questions.xlsx"),
        "output": Path("../data/processed/rag_unified/price_qa.jsonl"),
    },
    {
        "domain": "quality",
        "source_type": "qa",
        "input": Path("../data/uploads/conversations/qa_pairs_raw/quality_questions.xlsx"),
        "output": Path("../data/processed/rag_unified/quality_qa.jsonl"),
    },
]

SKU_SOURCE = {
    "domain": "spec",
    "source_type": "sku",
    "input": Path("../data/uploads/specs/sku_master.xlsx"),
    "output": Path("../data/processed/rag_unified/sku_master.jsonl"),
}


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").replace("\r\n", "\n").replace("\r", "\n").strip()


def _normalize_bool(value: Any) -> bool:
    text = _normalize_text(value).lower()
    return text in {"true", "1", "yes", "y", "是"}


def _split_semicolon(value: Any) -> List[str]:
    text = _normalize_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(";") if item.strip()]


def _read_xlsx_rows(file_path: Path, sheet_name: str) -> Iterable[Dict[str, Any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name} in {file_path}")

    worksheet = workbook[sheet_name]
    rows_iter = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    header_names = [
        _normalize_text(header) or f"column_{index + 1}"
        for index, header in enumerate(headers)
    ]

    rows: List[Dict[str, Any]] = []

    for excel_row_number, values in enumerate(rows_iter, start=2):
        row = {
            header_names[index]: values[index] if index < len(values) else None
            for index in range(len(header_names))
        }

        if any(_normalize_text(value) for value in row.values()):
            rows.append(
                {
                    **row,
                    "_excel_row_number": excel_row_number,
                }
            )

    return rows


def _build_qa_text(record: Dict[str, Any]) -> str:
    lines = [
        "【知识类型】QA",
        f"【业务域】{record['domain']}",
        f"【问题】{record['question_normalized'] or record['question_raw']}",
        f"【标准答案】{record['answer_standard']}",
        f"【意图子类】{record['intent_subtype']}",
        f"【关联SKU】{';'.join(record['related_sku_ids'])}",
        f"【必需字段】{';'.join(record['required_fields'])}",
        f"【是否需转人工】{str(record['handoff_required']).lower()}",
        f"【风险标记】{';'.join(record['risk_flags'])}",
        f"【审核状态】{record['verification_status']}",
        f"【答案来源】{record['answer_source']}",
    ]

    return "\n".join(line for line in lines if not line.endswith("】"))


def _build_sku_text(record: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "【知识类型】SKU",
            f"【业务域】{record['domain']}",
            f"【SKU】{record['sku_id']}",
            f"【产品名称】{record['product_name']}",
            f"【螺纹规格】{record['thread_spec']}",
            f"【杆长mm】{record['rod_length_mm']}",
            f"【球径mm】{record['ball_diameter_mm']}",
            f"【锥度比】{record['taper_ratio']}",
            f"【材质】{record['material']}",
            f"【表面处理】{record['surface_finish']}",
            f"【OEM对照号】{record['oem_reference_number']}",
            f"【起订量】{record['min_order_qty']}",
            f"【备货状态】{record['stock_status']}",
            f"【发货周期天】{record['lead_time_days']}",
        ]
    )


def _convert_qa_file(source: Dict[str, Any]) -> int:
    input_path: Path = source["input"]
    output_path: Path = source["output"]
    domain: str = source["domain"]
    source_type: str = source["source_type"]

    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    converted_count = 0

    with output_path.open("w", encoding="utf-8") as writer:
        for row in _read_xlsx_rows(input_path, sheet_name="qa_pairs"):
            answer_standard = _normalize_text(row.get("answer_standard"))
            question_raw = _normalize_text(row.get("question_raw"))
            question_normalized = _normalize_text(row.get("question_normalized"))

            if not answer_standard:
                continue

            record = {
                "text": "",
                "record_type": "qa",
                "domain": domain,
                "source_type": source_type,
                "qa_id": _normalize_text(row.get("qa_id")),
                "source_group_id": _normalize_text(row.get("source_group_id")),
                "primary_intent": _normalize_text(row.get("primary_intent")) or domain,
                "secondary_intents": _split_semicolon(row.get("secondary_intents")),
                "intent_subtype": _normalize_text(row.get("intent_subtype")),
                "question_raw": question_raw,
                "question_normalized": question_normalized,
                "answer_standard": answer_standard,
                "related_sku_ids": _split_semicolon(row.get("related_sku_ids")),
                "required_fields": _split_semicolon(row.get("required_fields")),
                "answer_source": _normalize_text(row.get("answer_source")),
                "handoff_required": _normalize_bool(row.get("handoff_required")),
                "risk_flags": _split_semicolon(row.get("risk_flags")),
                "verification_status": _normalize_text(row.get("verification_status")),
                "review_notes": _normalize_text(row.get("review_notes")),
                "source_file": str(input_path),
                "source_sheet": "qa_pairs",
                "excel_row_number": row.get("_excel_row_number"),
            }
            record["text"] = _build_qa_text(record)

            writer.write(json.dumps(record, ensure_ascii=False) + "\n")
            converted_count += 1

    return converted_count


def _convert_sku_file(source: Dict[str, Any]) -> int:
    input_path: Path = source["input"]
    output_path: Path = source["output"]
    domain: str = source["domain"]
    source_type: str = source["source_type"]

    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    converted_count = 0

    with output_path.open("w", encoding="utf-8") as writer:
        for row in _read_xlsx_rows(input_path, sheet_name="Sheet1"):
            sku_id = _normalize_text(row.get("SKU_ID"))
            product_name = _normalize_text(row.get("产品名称"))

            if not sku_id or not product_name:
                continue

            record = {
                "text": "",
                "record_type": "sku",
                "domain": domain,
                "source_type": source_type,
                "sku_id": sku_id,
                "product_name": product_name,
                "thread_spec": _normalize_text(row.get("螺纹规格")),
                "rod_length_mm": _normalize_text(row.get("杆长(mm)")),
                "ball_diameter_mm": _normalize_text(row.get("球径(mm)")),
                "taper_ratio": _normalize_text(row.get("锥度比")),
                "material": _normalize_text(row.get("材质")),
                "surface_finish": _normalize_text(row.get("表面处理")),
                "oem_reference_number": _normalize_text(row.get("OEM对照号")),
                "min_order_qty": _normalize_text(row.get("起订量(个)")),
                "stock_status": _normalize_text(row.get("备货状态")),
                "lead_time_days": _normalize_text(row.get("发货周期(天)")),
                "source_file": str(input_path),
                "source_sheet": "Sheet1",
                "excel_row_number": row.get("_excel_row_number"),
                "verification_status": "source_master",
            }
            record["text"] = _build_sku_text(record)

            writer.write(json.dumps(record, ensure_ascii=False) + "\n")
            converted_count += 1

    return converted_count


def build_manifest(results: List[Dict[str, Any]], output_dir: Path) -> None:
    manifest_path = output_dir / "manifest.json"

    total_count = sum(item["count"] for item in results)
    payload = {
        "total_count": total_count,
        "files": results,
        "notes": [
            "business_rules.md is not included because current PowerShell output shows mojibake and needs encoding verification.",
            "answer_raw is intentionally excluded from embedding text to avoid stale price/logistics claims entering automatic answers.",
            "answer_standard is used as the primary grounded answer field.",
        ],
    }

    manifest_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build unified JSONL files for rag_chunks_v1 ingestion."
    )
    parser.add_argument(
        "--output-dir",
        default="../data/processed/rag_unified",
        help="Output directory for unified JSONL files.",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    results: List[Dict[str, Any]] = []

    for source in QA_SOURCES:
        count = _convert_qa_file(source)
        results.append(
            {
                "domain": source["domain"],
                "source_type": source["source_type"],
                "input": str(source["input"]),
                "output": str(source["output"]),
                "count": count,
            }
        )
        print(f"[build] {source['domain']} qa count={count} output={source['output']}")

    sku_count = _convert_sku_file(SKU_SOURCE)
    results.append(
        {
            "domain": SKU_SOURCE["domain"],
            "source_type": SKU_SOURCE["source_type"],
            "input": str(SKU_SOURCE["input"]),
            "output": str(SKU_SOURCE["output"]),
            "count": sku_count,
        }
    )
    print(f"[build] sku count={sku_count} output={SKU_SOURCE['output']}")

    build_manifest(results, output_dir)

    print(f"[build] total_count={sum(item['count'] for item in results)}")
    print(f"[build] manifest={output_dir / 'manifest.json'}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
