"""Inspect Phase 3-I-I evaluation risk gate logic for remaining blockers."""

from __future__ import annotations

import json
import re
from pathlib import Path
from pprint import pprint
from typing import Any, Final

from openpyxl import load_workbook


BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

EVAL_SCRIPT: Final[Path] = BACKEND_ROOT / "scripts/check_phase3ii_real_llm_50_case_eval.py"
TEST_CASES_FILE: Final[Path] = PROJECT_ROOT / "data/evaluation/test_cases_draft.xlsx"
REPORT_FILE: Final[Path] = (
    PROJECT_ROOT / "logs/evaluation/phase3ii_real_llm_50_case_eval_report.json"
)

TARGET_CASE_IDS: Final[set[str]] = {
    "TC_SPEC_010",
    "TC_SPEC_015",
    "TC_SPEC_017",
}

RISK_LOGIC_MARKERS: Final[tuple[str, ...]] = (
    "risk case was not gated",
    "risk_gate",
    "risk_flags",
    "scenario_type",
    "expected_handoff",
    "handoff_required",
    "render_safety_blocked",
    "answer_handoff_required",
    "answer_safety_blocked",
    "BLOCKER",
)


def main() -> int:
    """Inspect eval risk gate logic and target case data."""

    print("=" * 80)
    print("inspecting Phase 3-I-I eval risk gate logic")

    errors: list[str] = []

    result = {
        "eval_script_result": inspect_eval_script(errors=errors),
        "test_case_rows": inspect_test_case_rows(errors=errors),
        "report_cases": inspect_report_cases(errors=errors),
        "errors": errors,
    }

    pprint(result)

    if errors:
        print("Phase 3-I-I eval risk gate logic inspection failed")
        return 1

    print("Phase 3-I-I eval risk gate logic inspection passed")
    return 0


def inspect_eval_script(
    *,
    errors: list[str],
) -> dict[str, Any]:
    """Inspect evaluator source logic around risk gate."""

    if not EVAL_SCRIPT.exists():
        errors.append(f"missing eval script: {EVAL_SCRIPT}")
        return {"exists": False, "path": str(EVAL_SCRIPT)}

    content = EVAL_SCRIPT.read_text(encoding="utf-8")

    return {
        "exists": True,
        "path": str(EVAL_SCRIPT.relative_to(BACKEND_ROOT)),
        "risk_related_lines": extract_lines(
            content=content,
            markers=RISK_LOGIC_MARKERS,
            limit=240,
        ),
        "functions": re.findall(
            r"^def ([a-zA-Z_][a-zA-Z0-9_]*)\(",
            content,
            flags=re.MULTILINE,
        ),
    }


def inspect_test_case_rows(
    *,
    errors: list[str],
) -> list[dict[str, Any]]:
    """Inspect workbook rows for target cases."""

    if not TEST_CASES_FILE.exists():
        errors.append(f"missing test cases file: {TEST_CASES_FILE}")
        return []

    workbook = load_workbook(TEST_CASES_FILE, data_only=True)
    sheet = workbook["test_cases"]

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]

    rows: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }

        case_id = str(row_data.get("case_id") or "")

        if case_id in TARGET_CASE_IDS:
            rows.append(row_data)

    return rows


def inspect_report_cases(
    *,
    errors: list[str],
) -> list[dict[str, Any]]:
    """Inspect latest report cases."""

    if not REPORT_FILE.exists():
        errors.append(f"missing report file: {REPORT_FILE}")
        return []

    report = json.loads(REPORT_FILE.read_text(encoding="utf-8"))
    results = report.get("results", [])

    if not isinstance(results, list):
        errors.append("report results must be list")
        return []

    selected: list[dict[str, Any]] = []

    for item in results:
        if not isinstance(item, dict):
            continue

        if str(item.get("case_id") or "") in TARGET_CASE_IDS:
            selected.append(
                {
                    "case_id": item.get("case_id"),
                    "category": item.get("category"),
                    "scenario_type": item.get("scenario_type"),
                    "expected_module": item.get("expected_module"),
                    "selected_module": item.get("selected_module"),
                    "answer_strategy_mode": item.get("answer_strategy_mode"),
                    "handoff_required": item.get("handoff_required"),
                    "answer_handoff_required": item.get("answer_handoff_required"),
                    "render_safety_blocked": item.get("render_safety_blocked"),
                    "failure_reasons": item.get("failure_reasons"),
                    "passed": item.get("passed"),
                    "final_response_preview": str(
                        item.get("final_response_preview") or ""
                    )[:500],
                }
            )

    return selected


def extract_lines(
    *,
    content: str,
    markers: tuple[str, ...],
    limit: int,
) -> list[str]:
    """Extract lines containing markers."""

    lines: list[str] = []

    for line_number, line in enumerate(content.splitlines(), start=1):
        lowered = line.lower()

        if any(marker.lower() in lowered for marker in markers):
            lines.append(f"{line_number}: {line.rstrip()}")

    return lines[:limit]


if __name__ == "__main__":
    raise SystemExit(main())