"""Validate sku_master.xlsx before database import.

This script only reads and validates the source Excel file. It does not connect
to PostgreSQL and does not write any data.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Final

from openpyxl import load_workbook  # type: ignore[import-untyped]

PROJECT_ROOT: Final[Path] = Path(__file__).resolve().parents[2]
SKU_FILE: Final[Path] = PROJECT_ROOT / "data" / "uploads" / "specs" / "sku_master.xlsx"

EXPECTED_HEADERS: Final[list[str]] = [
    "SKU_ID",
    "产品名称",
    "螺纹规格",
    "杆长(mm)",
    "球径(mm)",
    "锥度比",
    "材质",
    "表面处理",
    "OEM对照号",
    "起订量(个)",
    "备货状态",
    "发货周期(天)",
]

ALLOWED_THREAD_SPECS: Final[set[str]] = {
    "M8×1.25",
    "M10×1.5",
    "M12×1.25",
}

SKU_PATTERN: Final[re.Pattern[str]] = re.compile(r"^SKU\d{3}$")
OEM_PATTERN: Final[re.Pattern[str]] = re.compile(r"^\d{5}-\d{5}$")
TAPER_PATTERN: Final[re.Pattern[str]] = re.compile(r"^1:[1-9][0-9]*$")
THREAD_PATTERN: Final[re.Pattern[str]] = re.compile(
    r"^M(?P<diameter>\d+(?:\.\d+)?)×(?P<pitch>\d+(?:\.\d+)?)$"
)


@dataclass(frozen=True)
class ValidatedSkuRow:
    """Normalized SKU row after validation."""

    excel_row_number: int
    sku_id: str
    product_name: str
    thread_spec: str
    thread_diameter_mm: Decimal
    thread_pitch_mm: Decimal
    rod_length_mm: Decimal
    ball_diameter_mm: Decimal
    taper_ratio: str | None
    material: str
    surface_treatment: str
    oem_reference_number: str
    min_order_qty: int
    stock_status: str
    lead_time_days: int


def cell_to_text(value: object) -> str:
    """Convert an Excel cell value to stripped text."""

    if value is None:
        return ""

    return str(value).strip()


def parse_decimal(value: object, *, field_name: str, row_number: int) -> Decimal:
    """Parse a positive decimal value from an Excel cell."""

    text_value = cell_to_text(value)

    try:
        decimal_value = Decimal(text_value)
    except InvalidOperation as exc:
        raise ValueError(
            f"row {row_number}: {field_name} must be numeric, got {text_value!r}"
        ) from exc

    if decimal_value <= 0:
        raise ValueError(
            f"row {row_number}: {field_name} must be positive, got {text_value!r}"
        )

    return decimal_value


def parse_positive_int(value: object, *, field_name: str, row_number: int) -> int:
    """Parse a positive integer value from an Excel cell."""

    text_value = cell_to_text(value)

    try:
        decimal_value = Decimal(text_value)
    except InvalidOperation as exc:
        raise ValueError(
            f"row {row_number}: {field_name} must be an integer, got {text_value!r}"
        ) from exc

    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(
            f"row {row_number}: {field_name} must be an integer, got {text_value!r}"
        )

    int_value = int(decimal_value)

    if int_value <= 0:
        raise ValueError(
            f"row {row_number}: {field_name} must be positive, got {text_value!r}"
        )

    return int_value


def validate_header(actual_headers: list[str]) -> list[str]:
    """Validate exact Excel header order."""

    errors: list[str] = []

    if actual_headers != EXPECTED_HEADERS:
        errors.append("header mismatch")
        errors.append(f"expected: {EXPECTED_HEADERS}")
        errors.append(f"actual:   {actual_headers}")

    return errors


def validate_sku_row(row_number: int, values: list[object]) -> ValidatedSkuRow:
    """Validate and normalize one SKU row."""

    row = dict(zip(EXPECTED_HEADERS, values, strict=True))

    missing_fields = [
        field_name
        for field_name, raw_value in row.items()
        if cell_to_text(raw_value) == ""
    ]

    if missing_fields:
        raise ValueError(
            f"row {row_number}: blank fields: {', '.join(missing_fields)}"
        )

    sku_id = cell_to_text(row["SKU_ID"])
    if not SKU_PATTERN.fullmatch(sku_id):
        raise ValueError(f"row {row_number}: invalid SKU_ID {sku_id!r}")

    product_name = cell_to_text(row["产品名称"])

    thread_spec = cell_to_text(row["螺纹规格"])
    if thread_spec not in ALLOWED_THREAD_SPECS:
        raise ValueError(f"row {row_number}: unsupported thread_spec {thread_spec!r}")

    thread_match = THREAD_PATTERN.fullmatch(thread_spec)
    if thread_match is None:
        raise ValueError(f"row {row_number}: invalid thread_spec format {thread_spec!r}")

    thread_diameter_mm = Decimal(thread_match.group("diameter"))
    thread_pitch_mm = Decimal(thread_match.group("pitch"))

    rod_length_mm = parse_decimal(
        row["杆长(mm)"],
        field_name="杆长(mm)",
        row_number=row_number,
    )
    ball_diameter_mm = parse_decimal(
        row["球径(mm)"],
        field_name="球径(mm)",
        row_number=row_number,
    )

    raw_taper_ratio = cell_to_text(row["锥度比"])
    if raw_taper_ratio == "无":
        taper_ratio: str | None = None
    elif TAPER_PATTERN.fullmatch(raw_taper_ratio):
        taper_ratio = raw_taper_ratio
    else:
        raise ValueError(f"row {row_number}: invalid taper_ratio {raw_taper_ratio!r}")

    material = cell_to_text(row["材质"])
    surface_treatment = cell_to_text(row["表面处理"])

    oem_reference_number = cell_to_text(row["OEM对照号"])
    if not OEM_PATTERN.fullmatch(oem_reference_number):
        raise ValueError(
            f"row {row_number}: invalid OEM reference {oem_reference_number!r}"
        )

    min_order_qty = parse_positive_int(
        row["起订量(个)"],
        field_name="起订量(个)",
        row_number=row_number,
    )
    stock_status = cell_to_text(row["备货状态"])
    lead_time_days = parse_positive_int(
        row["发货周期(天)"],
        field_name="发货周期(天)",
        row_number=row_number,
    )

    return ValidatedSkuRow(
        excel_row_number=row_number,
        sku_id=sku_id,
        product_name=product_name,
        thread_spec=thread_spec,
        thread_diameter_mm=thread_diameter_mm,
        thread_pitch_mm=thread_pitch_mm,
        rod_length_mm=rod_length_mm,
        ball_diameter_mm=ball_diameter_mm,
        taper_ratio=taper_ratio,
        material=material,
        surface_treatment=surface_treatment,
        oem_reference_number=oem_reference_number,
        min_order_qty=min_order_qty,
        stock_status=stock_status,
        lead_time_days=lead_time_days,
    )


def validate_sku_file(path: Path) -> tuple[list[ValidatedSkuRow], list[str]]:
    """Validate the SKU workbook and return normalized rows plus errors."""

    errors: list[str] = []

    if not path.exists():
        return [], [f"file not found: {path}"]

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    raw_headers = [
        cell_to_text(cell.value)
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    ]

    errors.extend(validate_header(raw_headers))

    if errors:
        workbook.close()
        return [], errors

    validated_rows: list[ValidatedSkuRow] = []

    for row_number, excel_row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        values = list(excel_row)

        if all(cell_to_text(value) == "" for value in values):
            continue

        if len(values) != len(EXPECTED_HEADERS):
            errors.append(
                f"row {row_number}: expected {len(EXPECTED_HEADERS)} cells, "
                f"got {len(values)}"
            )
            continue

        try:
            validated_rows.append(validate_sku_row(row_number, values))
        except ValueError as exc:
            errors.append(str(exc))

    workbook.close()

    errors.extend(validate_cross_row_rules(validated_rows))

    return validated_rows, errors


def validate_cross_row_rules(rows: list[ValidatedSkuRow]) -> list[str]:
    """Validate uniqueness, continuity, and cross-row business constraints."""

    errors: list[str] = []

    if len(rows) != 100:
        errors.append(f"expected 100 SKU rows, got {len(rows)}")

    sku_counter = Counter(row.sku_id for row in rows)
    duplicate_skus = sorted(
        sku_id for sku_id, count in sku_counter.items() if count > 1
    )
    if duplicate_skus:
        errors.append("duplicate SKU_ID: " + ", ".join(duplicate_skus))

    expected_sku_ids = {f"SKU{index:03d}" for index in range(1, 101)}
    actual_sku_ids = {row.sku_id for row in rows}

    missing_skus = sorted(expected_sku_ids - actual_sku_ids)
    extra_skus = sorted(actual_sku_ids - expected_sku_ids)

    if missing_skus:
        errors.append("missing SKU_ID: " + ", ".join(missing_skus))

    if extra_skus:
        errors.append("unexpected SKU_ID: " + ", ".join(extra_skus))

    product_counter = Counter(row.product_name for row in rows)
    duplicate_product_names = sorted(
        product_name
        for product_name, count in product_counter.items()
        if count > 1
    )
    if duplicate_product_names:
        errors.append(
            "duplicate product_name: " + ", ".join(duplicate_product_names)
        )

    oem_counter = Counter(row.oem_reference_number for row in rows)
    duplicate_oems = sorted(
        oem for oem, count in oem_counter.items() if count > 1
    )
    if duplicate_oems:
        errors.append("duplicate OEM reference: " + ", ".join(duplicate_oems))

    return errors


def print_summary(rows: list[ValidatedSkuRow]) -> None:
    """Print validation summary."""

    print("sku validation summary")
    print(f"file: {SKU_FILE}")
    print(f"validated_rows: {len(rows)}")

    print("\nthread_spec distribution:")
    for key, value in sorted(Counter(row.thread_spec for row in rows).items()):
        print(f"- {key}: {value}")

    print("\ntaper_ratio distribution:")
    taper_values = [
        row.taper_ratio if row.taper_ratio is not None else "NULL_FROM_无"
        for row in rows
    ]
    for key, value in sorted(Counter(taper_values).items()):
        print(f"- {key}: {value}")

    print("\nstock_status distribution:")
    for key, value in sorted(Counter(row.stock_status for row in rows).items()):
        print(f"- {key}: {value}")

    if rows:
        rod_lengths = [row.rod_length_mm for row in rows]
        ball_diameters = [row.ball_diameter_mm for row in rows]
        min_order_quantities = [row.min_order_qty for row in rows]
        lead_time_days = [row.lead_time_days for row in rows]

        print("\nnumeric ranges:")
        print(f"- rod_length_mm: {min(rod_lengths)} - {max(rod_lengths)}")
        print(f"- ball_diameter_mm: {min(ball_diameters)} - {max(ball_diameters)}")
        print(
            f"- min_order_qty: {min(min_order_quantities)} - "
            f"{max(min_order_quantities)}"
        )
        print(f"- lead_time_days: {min(lead_time_days)} - {max(lead_time_days)}")


def main() -> int:
    """Run SKU validation."""

    rows, errors = validate_sku_file(SKU_FILE)

    if errors:
        print("sku validation failed")
        for error in errors:
            print(f"- {error}")

        return 1

    print_summary(rows)
    print("\nsku validation passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())