# ruff: noqa: E402,I001
"""Check Phase 3-I-C logistics_questions.xlsx quality."""

from __future__ import annotations

import json
import re
from pathlib import Path
from pprint import pprint
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]


BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

LOGISTICS_FILE: Final[Path] = (
    PROJECT_ROOT
    / "data"
    / "uploads"
    / "conversations"
    / "qa_pairs_raw"
    / "logistics_questions.xlsx"
)
SKU_MASTER_FILE: Final[Path] = (
    PROJECT_ROOT / "data" / "uploads" / "specs" / "sku_master.xlsx"
)

EXPECTED_QA_COUNT: Final[int] = 50

REQUIRED_QA_COLUMNS: Final[set[str]] = {
    "qa_id",
    "source_group_id",
    "question_raw",
    "question_normalized",
    "answer_raw",
    "answer_standard",
    "primary_intent",
    "intent_subtype",
    "related_sku_ids",
    "risk_flags",
    "handoff_required",
    "verification_status",
}

ALLOWED_VERIFICATION_STATUS: Final[set[str]] = {
    "pending",
    "verified",
    "rejected",
}

ALLOWED_HANDOFF_VALUES: Final[set[str]] = {
    "true",
    "false",
}

HIGH_RISK_FLAGS: Final[set[str]] = {
    "business_policy_confirmation",
    "shipping_fee_confirmation",
    "free_shipping_commitment",
    "delivery_time_commitment",
    "same_day_shipping_commitment",
    "carrier_commitment",
    "compensation_claim",
    "aftersale_commitment",
    "complaint",
    "data_conflict",
    "unsupported_claim",
}

COMMITMENT_FRAGMENTS: Final[tuple[str, ...]] = (
    "一定包邮",
    "保证包邮",
    "今天一定发",
    "当天一定发",
    "明天一定到",
    "三天必到",
    "保证到货",
    "保证不延误",
    "固定运费",
    "一定赔",
    "一定补发",
    "退货运费一定我们出",
)


def check_logistics_questions_file() -> bool:
    """Check logistics_questions.xlsx."""

    print("=" * 80)
    print("checking logistics_questions.xlsx")

    errors: list[str] = []
    warnings: list[str] = []

    if not LOGISTICS_FILE.exists():
        errors.append(f"logistics file not found: {LOGISTICS_FILE}")

    if not SKU_MASTER_FILE.exists():
        errors.append(f"sku master file not found: {SKU_MASTER_FILE}")

    if errors:
        pprint({"errors": errors, "warnings": warnings})
        return False

    sku_ids = load_sku_ids(SKU_MASTER_FILE)

    workbook = load_workbook(LOGISTICS_FILE, data_only=True)
    sheet_names = set(workbook.sheetnames)

    if "qa_pairs" not in sheet_names:
        errors.append("missing worksheet: qa_pairs")

    if "label_dictionary" not in sheet_names:
        errors.append("missing worksheet: label_dictionary")

    if errors:
        pprint({"errors": errors, "warnings": warnings})
        return False

    sheet = workbook["qa_pairs"]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        errors.append("qa_pairs worksheet is empty")
        pprint({"errors": errors, "warnings": warnings})
        return False

    headers = [normalize_header(value) for value in rows[0]]
    header_set = set(headers)

    missing_columns = sorted(REQUIRED_QA_COLUMNS - header_set)
    extra_columns = sorted(header_set - REQUIRED_QA_COLUMNS)

    if missing_columns:
        errors.append(f"missing required columns: {missing_columns}")

    records = [
        row_to_record(headers, row)
        for row in rows[1:]
        if not row_is_empty(row)
    ]

    if len(records) != EXPECTED_QA_COUNT:
        errors.append(
            f"qa record count must be {EXPECTED_QA_COUNT}, got {len(records)}"
        )

    qa_ids = [str(record.get("qa_id") or "").strip() for record in records]
    duplicated_qa_ids = sorted(find_duplicates(qa_ids))

    if duplicated_qa_ids:
        errors.append(f"duplicated qa_id values: {duplicated_qa_ids}")

    expected_qa_ids = [f"LOGI{index:04d}" for index in range(1, len(records) + 1)]

    if qa_ids != expected_qa_ids:
        errors.append(
            "qa_id must be continuous and ordered as "
            f"LOGI0001-LOGI{len(records):04d}"
        )

    source_group_ids = [
        str(record.get("source_group_id") or "").strip()
        for record in records
    ]

    source_group_prefix_errors = [
        value
        for value in source_group_ids
        if value and not value.startswith("LOGI_SRC")
    ]

    if source_group_prefix_errors:
        errors.append(
            "source_group_id should use LOGI_SRC prefix, bad values: "
            f"{sorted(set(source_group_prefix_errors))[:10]}"
        )

    primary_intents = sorted(
        {
            str(record.get("primary_intent") or "").strip()
            for record in records
            if str(record.get("primary_intent") or "").strip()
        }
    )

    if primary_intents != ["logistics"]:
        errors.append(f"primary_intent must only be logistics, got {primary_intents}")

    verification_status_seen: set[str] = set()
    intent_subtypes_seen: set[str] = set()
    risk_flags_seen: set[str] = set()
    handoff_counts = {"true": 0, "false": 0, "other": 0}
    related_sku_errors: list[str] = []
    high_risk_without_handoff: list[str] = []
    commitment_answer_rows: list[str] = []

    for record in records:
        qa_id = str(record.get("qa_id") or "").strip()

        question_raw = str(record.get("question_raw") or "").strip()
        question_normalized = str(record.get("question_normalized") or "").strip()
        answer_raw = str(record.get("answer_raw") or "").strip()
        answer_standard = str(record.get("answer_standard") or "").strip()
        intent_subtype = str(record.get("intent_subtype") or "").strip()
        verification_status = str(record.get("verification_status") or "").strip()
        handoff_required = str(record.get("handoff_required") or "").strip().lower()
        risk_flags = parse_semicolon_values(record.get("risk_flags"))
        related_sku_ids = parse_semicolon_values(record.get("related_sku_ids"))

        if intent_subtype:
            intent_subtypes_seen.add(intent_subtype)

        if verification_status:
            verification_status_seen.add(verification_status)

        risk_flags_seen.update(risk_flags)

        if not question_raw:
            errors.append(f"{qa_id}: question_raw is empty")

        if not question_normalized:
            errors.append(f"{qa_id}: question_normalized is empty")

        if not answer_raw:
            warnings.append(f"{qa_id}: answer_raw is empty")

        if not answer_standard:
            errors.append(f"{qa_id}: answer_standard is empty")

        if verification_status not in ALLOWED_VERIFICATION_STATUS:
            errors.append(
                f"{qa_id}: invalid verification_status={verification_status!r}"
            )

        if handoff_required in ALLOWED_HANDOFF_VALUES:
            handoff_counts[handoff_required] += 1
        else:
            handoff_counts["other"] += 1
            errors.append(f"{qa_id}: invalid handoff_required={handoff_required!r}")

        for sku_id in related_sku_ids:
            if sku_id and sku_id not in sku_ids:
                related_sku_errors.append(f"{qa_id}:{sku_id}")

        if risk_flags & HIGH_RISK_FLAGS and handoff_required != "true":
            high_risk_without_handoff.append(qa_id)

        if contains_commitment_fragment(answer_standard):
            commitment_answer_rows.append(qa_id)

    if related_sku_errors:
        errors.append(
            "related_sku_ids contain unknown SKU IDs: "
            f"{related_sku_errors[:20]}"
        )

    if high_risk_without_handoff:
        warnings.append(
            "high-risk logistics rows without handoff_required=true: "
            f"{high_risk_without_handoff}"
        )

    if commitment_answer_rows:
        warnings.append(
            "answer_standard contains logistics commitment fragments, "
            f"manual review required: {commitment_answer_rows}"
        )

    result = {
        "file": str(LOGISTICS_FILE),
        "sheet_names": workbook.sheetnames,
        "record_count": len(records),
        "expected_record_count": EXPECTED_QA_COUNT,
        "column_count": len(headers),
        "missing_columns": missing_columns,
        "extra_columns": extra_columns,
        "qa_id_first": qa_ids[0] if qa_ids else None,
        "qa_id_last": qa_ids[-1] if qa_ids else None,
        "primary_intents": primary_intents,
        "intent_subtypes_seen": sorted(intent_subtypes_seen),
        "verification_status_seen": sorted(verification_status_seen),
        "risk_flags_seen": sorted(risk_flags_seen),
        "handoff_counts": handoff_counts,
        "sku_master_count": len(sku_ids),
        "errors": errors,
        "warnings": warnings,
    }

    pprint(result)

    output_file = (
        PROJECT_ROOT
        / "data"
        / "parsed"
        / "logistics"
        / "logistics_questions_check_result.json"
    )
    output_file.parent.mkdir(parents=True, exist_ok=True)
    output_file.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if errors:
        print("logistics_questions.xlsx check failed")
        return False

    print("logistics_questions.xlsx check passed")
    return True


def load_sku_ids(
    sku_master_file: Path,
) -> set[str]:
    """Load SKU_ID values from sku_master.xlsx."""

    workbook = load_workbook(sku_master_file, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return set()

    headers = [normalize_header(value) for value in rows[0]]

    try:
        sku_id_index = headers.index("SKU_ID")
    except ValueError:
        sku_id_index = headers.index("sku_id")

    sku_ids: set[str] = set()

    for row in rows[1:]:
        if row_is_empty(row):
            continue

        value = row[sku_id_index] if sku_id_index < len(row) else None
        sku_id = str(value or "").strip()

        if sku_id:
            sku_ids.add(sku_id)

    return sku_ids


def row_to_record(
    headers: list[str],
    row: tuple[Any, ...],
) -> dict[str, Any]:
    """Convert row to record."""

    record: dict[str, Any] = {}

    for index, header in enumerate(headers):
        if not header:
            continue

        record[header] = row[index] if index < len(row) else None

    return record


def normalize_header(value: object) -> str:
    """Normalize header value."""

    return str(value or "").strip()


def row_is_empty(row: tuple[Any, ...]) -> bool:
    """Return whether a row is empty."""

    return all(value is None or str(value).strip() == "" for value in row)


def parse_semicolon_values(value: object) -> set[str]:
    """Parse semicolon-separated values."""

    text = str(value or "").strip()

    if not text:
        return set()

    return {
        item.strip()
        for item in re.split(r"[;；]", text)
        if item.strip()
    }


def find_duplicates(values: list[str]) -> set[str]:
    """Find duplicated non-empty values."""

    seen: set[str] = set()
    duplicates: set[str] = set()

    for value in values:
        if not value:
            continue

        if value in seen:
            duplicates.add(value)
        else:
            seen.add(value)

    return duplicates


def contains_commitment_fragment(text: str) -> bool:
    """Return whether answer contains forbidden logistics commitment fragment."""

    return any(fragment in text for fragment in COMMITMENT_FRAGMENTS)


def main() -> int:
    """Run check."""

    passed = check_logistics_questions_file()
    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())