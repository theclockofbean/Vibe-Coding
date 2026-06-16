# ruff: noqa: E402,I001
"""Inspect contracts needed for Phase 3-I-I 50-case real LLM evaluation."""

from __future__ import annotations

import inspect
import sys
from pathlib import Path
from pprint import pprint
from typing import Any, Final

import openpyxl  # type: ignore[import-untyped]

BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent

if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.agent import workflow as workflow_module
from app.agent.state import create_initial_agent_state
from app.agent.state import state_to_response_payload


TEST_CASE_CANDIDATES: Final[tuple[Path, ...]] = (
    PROJECT_ROOT / "test_cases_draft.xlsx",
    PROJECT_ROOT / "data/test_cases_draft.xlsx",
    PROJECT_ROOT / "backend/test_cases_draft.xlsx",
    PROJECT_ROOT / "backend/data/test_cases_draft.xlsx",
    PROJECT_ROOT / "docs/test_cases_draft.xlsx",
)

WORKBOOK_SHEET_CANDIDATES: Final[tuple[str, ...]] = (
    "test_cases",
    "Sheet1",
)


def main() -> int:
    """Inspect 50-case eval contracts."""

    print("=" * 80)
    print("inspecting Phase 3-I-I 50-case eval contract")

    errors: list[str] = []

    workbook_result = inspect_workbook(errors=errors)
    workflow_result = inspect_workflow_contract()
    state_result = inspect_state_contract()

    result = {
        "workbook_result": workbook_result,
        "workflow_result": workflow_result,
        "state_result": state_result,
        "errors": errors,
    }

    pprint(result)

    if errors:
        print("Phase 3-I-I 50-case eval contract inspection failed")
        return 1

    print("Phase 3-I-I 50-case eval contract inspection passed")
    return 0


def inspect_workbook(
    *,
    errors: list[str],
) -> dict[str, Any]:
    """Inspect available test case workbook."""

    found_path = find_test_case_file()

    if found_path is None:
        errors.append("test_cases_draft.xlsx not found in expected project paths")
        return {
            "searched_paths": [str(path) for path in TEST_CASE_CANDIDATES],
            "found": False,
        }

    workbook = openpyxl.load_workbook(found_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    target_sheet_name = select_sheet_name(sheet_names)

    if target_sheet_name is None:
        errors.append(
            "no usable worksheet found; expected one of "
            f"{list(WORKBOOK_SHEET_CANDIDATES)}"
        )
        return {
            "path": str(found_path),
            "found": True,
            "sheet_names": sheet_names,
            "target_sheet": None,
        }

    sheet = workbook[target_sheet_name]
    headers = [
        str(cell.value).strip()
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        if cell.value is not None and str(cell.value).strip()
    ]

    sample_rows = read_sample_rows(sheet=sheet, headers=headers, limit=5)
    row_count = max(sheet.max_row - 1, 0)

    return {
        "path": str(found_path),
        "found": True,
        "sheet_names": sheet_names,
        "target_sheet": target_sheet_name,
        "row_count_excluding_header": row_count,
        "headers": headers,
        "sample_rows": sample_rows,
    }


def find_test_case_file() -> Path | None:
    """Find test case workbook."""

    for path in TEST_CASE_CANDIDATES:
        if path.exists():
            return path

    matches = sorted(PROJECT_ROOT.rglob("test_cases_draft*.xlsx"))

    return matches[0] if matches else None


def select_sheet_name(
    sheet_names: list[str],
) -> str | None:
    """Select worksheet for test cases."""

    for candidate in WORKBOOK_SHEET_CANDIDATES:
        if candidate in sheet_names:
            return candidate

    return sheet_names[0] if sheet_names else None


def read_sample_rows(
    *,
    sheet: Any,
    headers: list[str],
    limit: int,
) -> list[dict[str, Any]]:
    """Read sample worksheet rows."""

    rows: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, max_row=limit + 1, values_only=True):
        row_dict: dict[str, Any] = {}

        for index, header in enumerate(headers):
            row_dict[header] = row[index] if index < len(row) else None

        rows.append(row_dict)

    return rows


def inspect_workflow_contract() -> dict[str, Any]:
    """Inspect workflow runner contracts."""

    result: dict[str, Any] = {
        "workflow_module": str(Path(workflow_module.__file__).relative_to(BACKEND_ROOT)),
        "has_run_agent_workflow": hasattr(workflow_module, "run_agent_workflow"),
        "has_build_agent_workflow": hasattr(workflow_module, "build_agent_workflow"),
        "has_AgentWorkflowNodes": hasattr(workflow_module, "AgentWorkflowNodes"),
    }

    if hasattr(workflow_module, "run_agent_workflow"):
        result["run_agent_workflow_signature"] = str(
            inspect.signature(workflow_module.run_agent_workflow)
        )

    if hasattr(workflow_module, "build_agent_workflow"):
        result["build_agent_workflow_signature"] = str(
            inspect.signature(workflow_module.build_agent_workflow)
        )

    if hasattr(workflow_module, "AgentWorkflowNodes"):
        result["AgentWorkflowNodes_signature"] = str(
            inspect.signature(workflow_module.AgentWorkflowNodes)
        )

    return result


def inspect_state_contract() -> dict[str, Any]:
    """Inspect state factory and response payload contracts."""

    return {
        "create_initial_agent_state_signature": str(
            inspect.signature(create_initial_agent_state)
        ),
        "state_to_response_payload_signature": str(
            inspect.signature(state_to_response_payload)
        ),
    }


if __name__ == "__main__":
    raise SystemExit(main())