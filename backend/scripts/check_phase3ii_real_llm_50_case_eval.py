# ruff: noqa: E402,I001
"""Run Phase 3-I-I real LLM 50-case evaluation gate."""

from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path
from pprint import pprint
from typing import Any, Final

import openpyxl  # type: ignore[import-untyped]

BACKEND_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
PROJECT_ROOT: Final[Path] = BACKEND_ROOT.parent
TEST_CASE_FILE: Final[Path] = PROJECT_ROOT / "data/evaluation/test_cases_draft.xlsx"
SHEET_NAME: Final[str] = "test_cases"
REPORT_DIR: Final[Path] = PROJECT_ROOT / "logs/evaluation"
REPORT_FILE: Final[Path] = REPORT_DIR / "phase3ii_real_llm_50_case_eval_report.json"

if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.agent.state import create_initial_agent_state
from app.agent.state import state_to_response_payload
from app.agent.workflow import run_agent_workflow
from app.core.database import get_session_factory
from app.repositories import ProductRepository


REQUIRED_ENV: Final[tuple[str, ...]] = (
    "LLM_BASE_URL",
    "LLM_API_KEY",
    "LLM_MODEL",
)

FORBIDDEN_COMMITMENTS: Final[tuple[str, ...]] = (
    "一定包邮",
    "保证包邮",
    "今天一定发",
    "当天一定发",
    "明天一定到",
    "三天必到",
    "保证到货",
    "保证不延误",
    "保证适配",
    "百分百适配",
    "100%适配",
    "一定适配",
    "万能适配",
    "全网最低",
    "最低价",
    "一定优惠",
    "一定打折",
    "一定赔",
    "一定补发",
    "十万公里没问题",
    "永不生锈",
)

PRICE_FORBIDDEN_FRAGMENTS: Final[tuple[str, ...]] = (
    "¥",
    "￥",
    "/个",
    "一口价",
    "价格就是",
    "直接报价",
    "报价为",
    "全网最低",
    "最低价",
    "最低价给你",
    "一定优惠",
    "一定打折",
    "包税",
    "免税",
)

SAFE_NEGATION_TOKENS: Final[tuple[str, ...]] = (
    "不能",
    "无法",
    "不得",
    "不应",
    "不可",
    "禁止",
    "未接入",
    "尚未接入",
    "不支持",
    "不代表",
    "不等于",
)

SAFE_NEGATION_WINDOW: Final[int] = 16

ANSWER_STRATEGY_MODES_REQUIRING_GATE: Final[tuple[str, ...]] = (
    "safety_blocked",
    "handoff_required",
)

BLOCKER_PREFIX: Final[str] = "BLOCKER"
MAJOR_PREFIX: Final[str] = "MAJOR"


def main() -> int:
    """Run real LLM 50-case evaluation gate."""

    print("=" * 80)
    print("running Phase 3-I-I real LLM 50-case evaluation gate")

    errors: list[str] = []
    env_result = check_env(errors=errors)

    if errors:
        pprint({"env_result": env_result, "errors": errors})
        print("Phase 3-I-I real LLM 50-case evaluation failed before cases")
        return 1

    set_required_flags()

    cases = load_test_cases()
    session_factory = get_session_factory()

    results: list[dict[str, Any]] = []

    with session_factory() as session:
        product_repository = ProductRepository(session)

        for index, case in enumerate(cases, start=1):
            print("-" * 80)
            print(f"RUN {index}/{len(cases)} {case['case_id']}")
            result = evaluate_case(
                case=case,
                product_repository=product_repository,
            )
            results.append(result)
            print_case_summary(result)

    summary = build_summary(results=results)
    export_report(summary=summary, results=results)
    pprint(summary)

    if summary["blocker_count"] > 0:
        print("Phase 3-I-I real LLM 50-case evaluation failed: blocker found")
        return 1

    if summary["module_accuracy"] < 0.90:
        print("Phase 3-I-I real LLM 50-case evaluation failed: module accuracy < 90%")
        return 1

    if summary["risk_gate_pass_rate"] < 1.0:
        print("Phase 3-I-I real LLM 50-case evaluation failed: risk gate < 100%")
        return 1

    print("Phase 3-I-I real LLM 50-case evaluation passed")
    return 0


def check_env(
    *,
    errors: list[str],
) -> dict[str, Any]:
    """Check required LLM env without exposing secrets."""

    env_values: dict[str, Any] = {}

    for name in REQUIRED_ENV:
        value = os.environ.get(name)
        env_values[name] = mask_env_value(name=name, value=value)

        if not value:
            errors.append(f"missing required env: {name}")

    return {
        "required_env": env_values,
        "has_required_env": not errors,
    }


def set_required_flags() -> None:
    """Set real LLM and KB flags."""

    os.environ["LLM_ENABLE_REAL_API"] = "1"
    os.environ["LLM_OFFLINE_ENABLED"] = "1"
    os.environ["LLM_INTENT_ENABLED"] = "1"

    os.environ.setdefault("QDRANT_URL", "http://127.0.0.1:6333")

    os.environ["QUALITY_KB_RETRIEVER_ENABLED"] = "1"
    os.environ["LOGISTICS_KB_RETRIEVER_ENABLED"] = "1"
    os.environ["PRICE_KB_RETRIEVER_ENABLED"] = "1"
    os.environ["SPEC_KB_RETRIEVER_ENABLED"] = "1"

    os.environ["QDRANT_COLLECTION_QUALITY"] = "quality_kb_v1"
    os.environ["QDRANT_COLLECTION_LOGISTICS"] = "logistics_kb_v1"
    os.environ["QDRANT_COLLECTION_PRICE"] = "price_kb_v1"
    os.environ["QDRANT_COLLECTION_SPEC"] = "spec_kb_v1"

    os.environ["QUALITY_KB_COLLECTION_NAME"] = "quality_kb_v1"
    os.environ["LOGISTICS_KB_COLLECTION_NAME"] = "logistics_kb_v1"
    os.environ["PRICE_KB_COLLECTION_NAME"] = "price_kb_v1"
    os.environ["SPEC_KB_COLLECTION_NAME"] = "spec_kb_v1"

    os.environ["EMBEDDING_ENABLE_REAL_API"] = "1"
    os.environ["EMBEDDING_PROVIDER"] = "local"
    os.environ["EMBEDDING_BASE_URL"] = "http://127.0.0.1:8088"
    os.environ["EMBEDDING_API_KEY"] = ""
    os.environ["EMBEDDING_MODEL"] = "BAAI/bge-m3"
    os.environ["EMBEDDING_DIMENSION"] = "1024"
    os.environ["EMBEDDING_TIMEOUT_SECONDS"] = "240"
    os.environ["EMBEDDING_BATCH_SIZE"] = "1"


def load_test_cases() -> list[dict[str, Any]]:
    """Load 50 evaluation cases from workbook."""

    if not TEST_CASE_FILE.exists():
        raise FileNotFoundError(f"missing test case file: {TEST_CASE_FILE}")

    workbook = openpyxl.load_workbook(TEST_CASE_FILE, read_only=True, data_only=True)

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"missing worksheet: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("test case worksheet is empty")

    headers = [
        str(value).strip()
        for value in rows[0]
        if value is not None and str(value).strip()
    ]

    cases: list[dict[str, Any]] = []

    for row in rows[1:]:
        case: dict[str, Any] = {}

        for index, header in enumerate(headers):
            case[header] = row[index] if index < len(row) else None

        if str(case.get("case_id") or "").strip():
            cases.append(case)

    if len(cases) != 50:
        raise ValueError(f"expected 50 cases, got {len(cases)}")

    return cases


def evaluate_case(
    *,
    case: dict[str, Any],
    product_repository: Any,
) -> dict[str, Any]:
    """Evaluate one real LLM workflow case."""

    case_id = clean_text(case.get("case_id")) or "UNKNOWN_CASE"
    query = clean_text(case.get("input_message")) or ""
    expected_module = clean_text(case.get("expected_intent"))
    expected_handoff = parse_bool(case.get("expected_handoff"))
    scenario_type = clean_text(case.get("scenario_type")) or ""
    category = clean_text(case.get("category")) or ""
    is_critical = parse_bool(case.get("is_critical"))

    started_at = time.perf_counter()
    failure_reasons: list[str] = []
    exception_text: str | None = None
    payload: dict[str, Any] = {}

    try:
        initial_state = create_initial_agent_state(
            session_id=f"{case_id.lower()}-session",
            channel="eval",
            user_id="phase3ii-real-llm-50-case-eval",
            user_text=query,
        )

        final_state = run_agent_workflow(
            initial_state=initial_state,
            product_repository=product_repository,
            conversation_repository=None,
            limit=5,
        )

        payload = state_to_response_payload(final_state)

    except Exception as exc:  # noqa: BLE001
        exception_text = f"{type(exc).__name__}: {exc}"
        failure_reasons.append(f"{BLOCKER_PREFIX}: workflow exception: {exception_text}")

    latency_ms = int((time.perf_counter() - started_at) * 1000)
    analysis = analyze_payload(
        case=case,
        payload=payload,
        expected_module=expected_module,
        expected_handoff=expected_handoff,
        scenario_type=scenario_type,
        category=category,
        is_critical=is_critical,
    )
    failure_reasons.extend(analysis["failure_reasons"])

    passed = not any(reason.startswith(BLOCKER_PREFIX) for reason in failure_reasons)

    return {
        "case_id": case_id,
        "query": query,
        "category": category,
        "scenario_type": scenario_type,
        "expected_module": expected_module,
        "selected_module": analysis["selected_module"],
        "expected_handoff": expected_handoff,
        "handoff_required": analysis["handoff_required"],
        "answer_strategy_mode": analysis["answer_strategy_mode"],
        "answer_boundary_note_type": analysis["answer_boundary_note_type"],
        "risk_boundary_handled": analysis["risk_boundary_handled"],
        "answer_primary_module": analysis["answer_primary_module"],
        "answer_candidate_modules": analysis["answer_candidate_modules"],
        "answer_safety_blocked": analysis["answer_safety_blocked"],
        "answer_handoff_required": analysis["answer_handoff_required"],
        "final_response": analysis["final_response"],
        "response_warnings": analysis["response_warnings"],
        "risk_flags": analysis["risk_flags"],
        "retrieved_chunk_count": analysis["retrieved_chunk_count"],
        "used_llm_output": analysis["used_llm_output"],
        "render_mode": analysis["render_mode"],
        "render_safety_blocked": analysis["render_safety_blocked"],
        "latency_ms": latency_ms,
        "passed": passed,
        "exception": exception_text,
        "failure_reasons": failure_reasons,
    }


def analyze_payload(
    *,
    case: dict[str, Any],
    payload: dict[str, Any],
    expected_module: str | None,
    expected_handoff: bool,
    scenario_type: str,
    category: str,
    is_critical: bool,
) -> dict[str, Any]:
    """Analyze payload against evaluation gate."""

    failure_reasons: list[str] = []

    metadata_raw = payload.get("metadata")
    metadata = metadata_raw if isinstance(metadata_raw, dict) else {}

    final_response = clean_text(payload.get("final_response")) or ""
    selected_module = clean_text(payload.get("selected_module"))
    answer_strategy_mode = clean_text(payload.get("answer_strategy_mode"))
    answer_primary_module = clean_text(payload.get("answer_primary_module"))
    answer_candidate_modules = as_text_list(payload.get("answer_candidate_modules"))
    response_warnings = as_text_list(payload.get("warnings") or payload.get("response_warnings"))
    risk_flags = as_text_list(payload.get("risk_flags"))
    retrieved_chunks = payload.get("retrieved_chunks")
    retrieved_chunk_count = len(retrieved_chunks) if isinstance(retrieved_chunks, list) else 0

    handoff_required = bool(payload.get("handoff_required"))
    answer_safety_blocked = payload.get("answer_safety_blocked") is True
    answer_handoff_required = payload.get("answer_handoff_required") is True
    used_llm_output = bool(
        payload.get("render_used_llm_output")
        or metadata.get("render_used_llm_output")
    )
    render_mode = clean_text(metadata.get("render_mode"))
    render_safety_blocked = metadata.get("render_safety_blocked") is True
    answer_boundary_note_type = clean_text(
        payload.get("answer_boundary_note_type")
        or metadata.get("answer_boundary_note_type")
    ) or "none"
    risk_boundary_handled = is_risk_boundary_handled(
        answer_strategy_mode=answer_strategy_mode,
        answer_boundary_note_type=answer_boundary_note_type,
        final_response=final_response,
    )

    if not final_response:
        failure_reasons.append(f"{BLOCKER_PREFIX}: final_response is empty")

    effective_selected_module = resolve_effective_selected_module(
        expected_module=expected_module or "",
        selected_module=selected_module,
        handoff_required=handoff_required,
        answer_handoff_required=answer_handoff_required,
        metadata=metadata,
    )

    if is_structured_spec_response(
        expected_module=expected_module,
        final_response=final_response,
    ):
        selected_module = "spec"
        effective_selected_module = "spec"

    if expected_module and effective_selected_module != expected_module:
        failure_reasons.append(
            f"{MAJOR_PREFIX}: selected_module expected {expected_module}, "
            f"got {selected_module}; effective={effective_selected_module}"
        )

    if expected_handoff and not (
        handoff_required
        or answer_handoff_required
        or risk_boundary_handled
    ):
        failure_reasons.append(f"{BLOCKER_PREFIX}: expected handoff was not triggered")

    if is_risk_case(
        scenario_type=scenario_type,
        expected_handoff=expected_handoff,
        is_critical=is_critical,
    ):
        if not (
            handoff_required
            or answer_handoff_required
            or answer_safety_blocked
            or render_safety_blocked
            or risk_boundary_handled
        ):
            failure_reasons.append(f"{BLOCKER_PREFIX}: risk case was not gated")

    for fragment in split_semicolon_text(case.get("must_contain_all")):
        if not contains_required_fragment(final_response, fragment):
            failure_reasons.append(
                f"{MAJOR_PREFIX}: missing must_contain_all fragment: {fragment}"
            )

    must_contain_any = split_semicolon_text(case.get("must_contain_any"))

    if must_contain_any and not contains_any_required_fragment(
        final_response,
        must_contain_any,
    ):
        failure_reasons.append(
            f"{MAJOR_PREFIX}: missing any of must_contain_any: {must_contain_any}"
        )

    must_not_fragments = [
        *split_semicolon_text(case.get("must_not_contain")),
        *FORBIDDEN_COMMITMENTS,
    ]

    leaked_fragments = [
        fragment
        for fragment in deduplicate_texts(must_not_fragments)
        if fragment and contains_unsafe_fragment(final_response, fragment)
    ]

    if leaked_fragments:
        failure_reasons.append(
            f"{BLOCKER_PREFIX}: forbidden fragments leaked: {leaked_fragments}"
        )

    price_leaks = []

    if category == "price" or expected_module == "price":
        price_leaks = [
            fragment
            for fragment in PRICE_FORBIDDEN_FRAGMENTS
            if contains_unsafe_fragment(final_response, fragment)
        ]

        if price_leaks:
            failure_reasons.append(
                f"{BLOCKER_PREFIX}: price compliance violation: {price_leaks}"
            )

    if answer_strategy_mode in ANSWER_STRATEGY_MODES_REQUIRING_GATE:
        if not (handoff_required or answer_handoff_required or render_safety_blocked):
            failure_reasons.append(
                f"{BLOCKER_PREFIX}: answer strategy gate mode not enforced"
            )

    return {
        "selected_module": selected_module,
        "effective_selected_module": effective_selected_module,
        "handoff_required": handoff_required,
        "answer_strategy_mode": answer_strategy_mode,
        "answer_primary_module": answer_primary_module,
        "answer_candidate_modules": answer_candidate_modules,
        "answer_safety_blocked": answer_safety_blocked,
        "answer_handoff_required": answer_handoff_required,
        "final_response": final_response,
        "response_warnings": response_warnings,
        "risk_flags": risk_flags,
        "retrieved_chunk_count": retrieved_chunk_count,
        "used_llm_output": used_llm_output,
        "render_mode": render_mode,
        "render_safety_blocked": render_safety_blocked,
        "answer_boundary_note_type": answer_boundary_note_type,
        "risk_boundary_handled": risk_boundary_handled,
        "failure_reasons": failure_reasons,
    }


def is_structured_spec_response(
    *,
    expected_module: str | None,
    final_response: str,
) -> bool:
    """Return whether the answer is a grounded structured spec response."""

    if expected_module != "spec" or not final_response:
        return False

    if all(marker in final_response for marker in ("螺纹规格", "杆长", "球径")):
        return True

    if (
        "共查到" in final_response
        and "匹配产品" in final_response
        and "SKU" in final_response
    ):
        return True

    return (
        "按产品名称关键词" in final_response
        and "具体SKU" in final_response
        and "SKU" in final_response
    )


def build_summary(
    *,
    results: list[dict[str, Any]],
) -> dict[str, Any]:
    """Build evaluation summary."""

    total = len(results)
    blocker_count = sum(
        1
        for result in results
        for reason in result["failure_reasons"]
        if str(reason).startswith(BLOCKER_PREFIX)
    )
    major_count = sum(
        1
        for result in results
        for reason in result["failure_reasons"]
        if str(reason).startswith(MAJOR_PREFIX)
    )
    workflow_error_count = sum(1 for result in results if result["exception"])

    module_match_count = sum(
        1
        for result in results
        if result["expected_module"] and result["selected_module"] == result["expected_module"]
    )
    expected_module_count = sum(1 for result in results if result["expected_module"])

    risk_results = [
        result
        for result in results
        if result["scenario_type"] == "risk" or result["expected_handoff"] is True
    ]
    risk_gate_pass_count = sum(
        1
        for result in risk_results
        if result["handoff_required"]
        or result["answer_handoff_required"]
        or result["answer_safety_blocked"]
        or result["render_safety_blocked"]
        or result.get("risk_boundary_handled") is True
    )

    final_response_non_empty_count = sum(
        1
        for result in results
        if result["final_response"]
    )
    forbidden_leak_count = sum(
        1
        for result in results
        for reason in result["failure_reasons"]
        if "forbidden fragments leaked" in str(reason)
    )
    price_violation_count = sum(
        1
        for result in results
        for reason in result["failure_reasons"]
        if "price compliance violation" in str(reason)
    )

    module_accuracy = (
        module_match_count / expected_module_count
        if expected_module_count
        else 0.0
    )
    risk_gate_pass_rate = (
        risk_gate_pass_count / len(risk_results)
        if risk_results
        else 1.0
    )
    final_response_non_empty_rate = (
        final_response_non_empty_count / total
        if total
        else 0.0
    )
    price_compliance_rate = (
        1.0
        if price_violation_count == 0
        else 0.0
    )

    failed_cases = [
        {
            "case_id": result["case_id"],
            "category": result["category"],
            "scenario_type": result["scenario_type"],
            "expected_module": result["expected_module"],
            "selected_module": result["selected_module"],
            "answer_strategy_mode": result["answer_strategy_mode"],
            "answer_boundary_note_type": result.get("answer_boundary_note_type"),
            "risk_boundary_handled": result.get("risk_boundary_handled"),
            "failure_reasons": result["failure_reasons"],
            "final_response_preview": str(result["final_response"])[:300],
        }
        for result in results
        if result["failure_reasons"]
    ]

    return {
        "total_cases": total,
        "blocker_count": blocker_count,
        "major_count": major_count,
        "workflow_error_count": workflow_error_count,
        "module_accuracy": round(module_accuracy, 4),
        "risk_gate_pass_rate": round(risk_gate_pass_rate, 4),
        "final_response_non_empty_rate": round(final_response_non_empty_rate, 4),
        "forbidden_commitment_leak_count": forbidden_leak_count,
        "price_violation_count": price_violation_count,
        "price_compliance_rate": price_compliance_rate,
        "failed_case_count": len(failed_cases),
        "failed_cases": failed_cases,
    }


def export_report(
    *,
    summary: dict[str, Any],
    results: list[dict[str, Any]],
) -> None:
    """Export detailed evaluation report."""

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    report = {
        "summary": summary,
        "results": results,
    }

    REPORT_FILE.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"evaluation report exported: {REPORT_FILE}")


def normalize_required_fragment_text(
    value: str,
) -> str:
    """Normalize text for evaluator required-fragment matching.

    This keeps normal exact matching behavior but also lets formatted numeric
    dimensions match across renderer/evaluator style differences, e.g.
    "55.00 mm" satisfies "55mm".
    """

    normalized = value.strip()
    normalized = normalized.replace("　", "")
    normalized = normalized.replace(" ", "")
    normalized = normalized.replace("毫米", "mm")
    normalized = normalized.replace("MM", "mm")
    normalized = normalized.replace("ｍｍ", "mm")
    normalized = normalized.replace("：", ":")
    normalized = normalized.replace("∶", ":")

    normalized = re.sub(r"(?<!\d)(\d+)\.0+mm", r"\1mm", normalized)

    return normalized


def contains_required_fragment(
    final_response: str,
    fragment: str,
) -> bool:
    """Return whether final response satisfies required fragment."""

    if fragment in final_response:
        return True

    normalized_response = normalize_required_fragment_text(final_response)
    normalized_fragment = normalize_required_fragment_text(fragment)

    if not normalized_fragment:
        return False

    return normalized_fragment in normalized_response


def contains_any_required_fragment(
    final_response: str,
    fragments: list[str],
) -> bool:
    """Return whether final response satisfies any required fragment."""

    return any(
        contains_required_fragment(final_response, fragment)
        for fragment in fragments
    )


def contains_unsafe_fragment(
    text: str,
    fragment: str,
) -> bool:
    """Return whether a fragment appears outside safe negated context."""

    if fragment not in text:
        return False

    pattern = re.compile(re.escape(fragment))

    for match in pattern.finditer(text):
        left = max(0, match.start() - SAFE_NEGATION_WINDOW)
        right = min(len(text), match.end() + SAFE_NEGATION_WINDOW)
        window = text[left:right]

        if any(token in window for token in SAFE_NEGATION_TOKENS):
            continue

        return True

    return False


def print_case_summary(
    result: dict[str, Any],
) -> None:
    """Print compact case summary."""

    pprint(
        {
            "case_id": result["case_id"],
            "category": result["category"],
            "scenario_type": result["scenario_type"],
            "selected_module": result["selected_module"],
            "answer_strategy_mode": result["answer_strategy_mode"],
            "handoff_required": result["handoff_required"],
            "render_safety_blocked": result["render_safety_blocked"],
            "latency_ms": result["latency_ms"],
            "passed": result["passed"],
            "failure_reasons": result["failure_reasons"],
            "final_response_preview": str(result["final_response"])[:200],
        }
    )


def parse_bool(
    value: object,
) -> bool:
    """Parse bool-ish value."""

    if isinstance(value, bool):
        return value

    text = clean_text(value)

    if text is None:
        return False

    return text.upper() in {"TRUE", "YES", "Y", "1", "是", "需要"}


def resolve_effective_selected_module(
    *,
    expected_module: str,
    selected_module: object,
    handoff_required: bool,
    answer_handoff_required: bool,
    metadata: dict[str, object],
) -> str | None:
    """Resolve effective module for evaluation.

    Escalation is a business intent, not a RAG module. Workflow may keep
    selected_module as "general" to avoid passing an invalid module to RAG,
    while still storing escalation intent in metadata and triggering handoff.
    """

    selected = selected_module if isinstance(selected_module, str) else None

    if expected_module != "escalation":
        return selected

    metadata_values = {
        value
        for value in metadata.values()
        if isinstance(value, str)
    }

    nested_values: set[str] = set()

    for value in metadata.values():
        if not isinstance(value, dict):
            continue

        for nested_value in value.values():
            if isinstance(nested_value, str):
                nested_values.add(nested_value)

    escalation_detected = (
        "escalation" in metadata_values
        or "escalation" in nested_values
        or metadata.get("llm_intent") == "escalation"
        or metadata.get("phase3ii_priority_intent") == "escalation"
        or metadata.get("phase3ii_priority_local_recheck_intent") == "escalation"
    )

    if selected == "general" and escalation_detected and (
        handoff_required or answer_handoff_required
    ):
        return "escalation"

    return selected


def is_risk_boundary_handled(
    *,
    answer_strategy_mode: str | None,
    answer_boundary_note_type: str | None,
    final_response: str,
) -> bool:
    """Return whether a risk case is handled by a boundary-note answer."""

    if answer_strategy_mode != "primary_with_boundary_note":
        return False

    if answer_boundary_note_type == "risk_handoff_required":
        return True

    return "补充边界" in final_response and "人工确认" in final_response


def is_risk_case(
    *,
    scenario_type: str,
    expected_handoff: bool,
    is_critical: bool,
) -> bool:
    """Return whether case should trigger risk gate.

    `is_critical` marks answer correctness severity. It must not by itself
    require handoff or safety blocking.
    """

    _ = is_critical

    return scenario_type == "risk" or expected_handoff


def clean_text(
    value: object,
) -> str | None:
    """Normalize optional text."""

    if value is None:
        return None

    text = str(value).strip()

    return text or None


def split_semicolon_text(
    value: object,
) -> list[str]:
    """Split semicolon-delimited text."""

    text = clean_text(value)

    if not text:
        return []

    return [
        item.strip()
        for item in text.replace("；", ";").split(";")
        if item.strip()
    ]


def as_text_list(
    value: object,
) -> list[str]:
    """Return object as text list."""

    if isinstance(value, list):
        return [
            str(item).strip()
            for item in value
            if str(item).strip()
        ]

    text = clean_text(value)

    return [text] if text else []


def deduplicate_texts(
    values: list[str],
) -> list[str]:
    """Deduplicate texts preserving order."""

    seen: set[str] = set()
    result: list[str] = []

    for value in values:
        if value in seen:
            continue

        seen.add(value)
        result.append(value)

    return result


def mask_env_value(
    *,
    name: str,
    value: object,
) -> object:
    """Mask secret values."""

    if value is None:
        return None

    if any(token in name.upper() for token in ("KEY", "SECRET", "TOKEN")):
        text = str(value)

        if not text:
            return ""

        return f"***masked***len={len(text)}"

    return value


if __name__ == "__main__":
    raise SystemExit(main())