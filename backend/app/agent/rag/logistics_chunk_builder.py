"""Build Logistics KB chunks from logistics QA spreadsheet."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]


LOGISTICS_MODULE: Final[str] = "logistics"
LOGISTICS_SOURCE_TYPE: Final[str] = "qa_pair"
LOGISTICS_SOURCE_NAME: Final[str] = "logistics_questions.xlsx"
LOGISTICS_COLLECTION_NAME: Final[str] = "logistics_kb_v1"

HIGH_RISK_FLAGS: Final[set[str]] = {
    "business_policy_confirmation",
    "shipping_fee_confirmation",
    "free_shipping_commitment",
    "delivery_time_commitment",
    "same_day_shipping_commitment",
    "carrier_commitment",
    "compensation_claim",
    "aftersale_commitment",
    "complaint",
    "data_conflict",
    "unsupported_claim",
}

COMMITMENT_FRAGMENTS: Final[tuple[str, ...]] = (
    "一定包邮",
    "保证包邮",
    "今天一定发",
    "当天一定发",
    "明天一定到",
    "三天必到",
    "保证到货",
    "保证不延误",
    "固定运费",
    "一定赔",
    "一定补发",
    "退货运费一定我们出",
)

MEDIUM_RISK_TERMS: Final[tuple[str, ...]] = (
    "发货",
    "到货",
    "运费",
    "快递",
    "物流",
    "时效",
    "偏远",
    "包邮",
)


@dataclass(frozen=True)
class LogisticsQARecord:
    """Raw logistics QA record."""

    qa_id: str
    source_group_id: str
    question_raw: str
    question_normalized: str
    answer_raw: str
    answer_standard: str
    primary_intent: str
    intent_subtype: str
    related_sku_ids: tuple[str, ...]
    secondary_intents: tuple[str, ...]
    risk_flags: tuple[str, ...]
    handoff_required: bool
    verification_status: str
    source_row_index: int


@dataclass(frozen=True)
class LogisticsKBChunk:
    """Logistics KB chunk for PostgreSQL metadata and Qdrant payload."""

    chunk_id: str
    qa_id: str
    module: str
    source_type: str
    source_name: str
    source_row_index: int
    source_group_id: str
    content: str
    summary: str
    question: str
    answer: str
    intent_scope: tuple[str, ...]
    sku_scope: tuple[str, ...]
    risk_flags: tuple[str, ...]
    risk_level: str
    handoff_required: bool
    verification_status: str
    is_active: bool
    allow_answer_reference: bool
    allow_commitment_reference: bool
    qdrant_collection_name: str
    metadata: dict[str, str]


def load_logistics_qa_records(
    file_path: Path,
) -> list[LogisticsQARecord]:
    """Load logistics QA records from xlsx."""

    workbook = load_workbook(file_path, data_only=True)

    if "qa_pairs" not in workbook.sheetnames:
        raise ValueError("missing worksheet: qa_pairs")

    sheet = workbook["qa_pairs"]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_text(value) for value in rows[0]]
    records: list[LogisticsQARecord] = []

    for row_index, row in enumerate(rows[1:], start=2):
        if row_is_empty(row):
            continue

        row_dict = row_to_dict(headers, row)

        records.append(
            LogisticsQARecord(
                qa_id=get_required_text(row_dict, "qa_id", row_index),
                source_group_id=get_text(row_dict, "source_group_id"),
                question_raw=get_required_text(
                    row_dict,
                    "question_raw",
                    row_index,
                ),
                question_normalized=get_required_text(
                    row_dict,
                    "question_normalized",
                    row_index,
                ),
                answer_raw=get_text(row_dict, "answer_raw"),
                answer_standard=get_required_text(
                    row_dict,
                    "answer_standard",
                    row_index,
                ),
                primary_intent=get_required_text(
                    row_dict,
                    "primary_intent",
                    row_index,
                ),
                intent_subtype=get_text(row_dict, "intent_subtype"),
                related_sku_ids=parse_semicolon_tuple(
                    row_dict.get("related_sku_ids")
                ),
                secondary_intents=parse_semicolon_tuple(
                    row_dict.get("secondary_intents")
                ),
                risk_flags=parse_semicolon_tuple(row_dict.get("risk_flags")),
                handoff_required=parse_bool(row_dict.get("handoff_required")),
                verification_status=get_required_text(
                    row_dict,
                    "verification_status",
                    row_index,
                ),
                source_row_index=row_index,
            )
        )

    return records


def build_logistics_kb_chunks(
    records: list[LogisticsQARecord],
) -> list[LogisticsKBChunk]:
    """Build Logistics KB chunks from records."""

    return [build_logistics_kb_chunk(record) for record in records]


def build_logistics_kb_chunk(
    record: LogisticsQARecord,
) -> LogisticsKBChunk:
    """Build one Logistics KB chunk."""

    risk_flags = tuple(sorted(set(record.risk_flags)))
    risk_level = infer_risk_level(
        answer=record.answer_standard,
        risk_flags=set(risk_flags),
        handoff_required=record.handoff_required,
    )
    is_active = record.verification_status != "rejected"
    allow_answer_reference = is_active and bool(record.answer_standard.strip())

    intent_scope = tuple(
        item
        for item in (
            record.primary_intent,
            record.intent_subtype,
            *record.secondary_intents,
        )
        if item
    )

    content = build_chunk_content(record)

    return LogisticsKBChunk(
        chunk_id=f"logistics_qa_{record.qa_id.lower()}",
        qa_id=record.qa_id,
        module=LOGISTICS_MODULE,
        source_type=LOGISTICS_SOURCE_TYPE,
        source_name=LOGISTICS_SOURCE_NAME,
        source_row_index=record.source_row_index,
        source_group_id=record.source_group_id,
        content=content,
        summary=build_summary(record),
        question=record.question_normalized,
        answer=record.answer_standard,
        intent_scope=intent_scope,
        sku_scope=record.related_sku_ids,
        risk_flags=risk_flags,
        risk_level=risk_level,
        handoff_required=record.handoff_required,
        verification_status=record.verification_status,
        is_active=is_active,
        allow_answer_reference=allow_answer_reference,
        allow_commitment_reference=False,
        qdrant_collection_name=LOGISTICS_COLLECTION_NAME,
        metadata={
            "qa_id": record.qa_id,
            "source_group_id": record.source_group_id,
            "intent_subtype": record.intent_subtype,
            "verification_status": record.verification_status,
            "risk_level": risk_level,
        },
    )


def build_chunk_content(
    record: LogisticsQARecord,
) -> str:
    """Build retrievable chunk content."""

    parts = [
        f"问题：{record.question_normalized}",
        f"标准回答：{record.answer_standard}",
    ]

    if record.related_sku_ids:
        parts.append(f"关联 SKU：{';'.join(record.related_sku_ids)}")

    if record.risk_flags:
        parts.append(f"风险标记：{';'.join(record.risk_flags)}")

    parts.append("物流回答边界：不得承诺包邮、固定运费、确定到货时间或赔付。")

    return "\n".join(parts)


def build_summary(
    record: LogisticsQARecord,
) -> str:
    """Build chunk summary."""

    subtype = record.intent_subtype or "general"
    return f"{record.qa_id}｜物流问答｜{subtype}｜{record.question_normalized}"


def infer_risk_level(
    *,
    answer: str,
    risk_flags: set[str],
    handoff_required: bool,
) -> str:
    """Infer logistics risk level."""

    if handoff_required:
        return "high"

    if risk_flags & HIGH_RISK_FLAGS:
        return "high"

    if contains_commitment_fragment(answer):
        return "high"

    if any(term in answer for term in MEDIUM_RISK_TERMS):
        return "medium"

    return "low"


def contains_commitment_fragment(
    text: str,
) -> bool:
    """Return whether text contains forbidden commitment fragment."""

    return any(fragment in text for fragment in COMMITMENT_FRAGMENTS)


def row_to_dict(
    headers: list[str],
    row: tuple[Any, ...],
) -> dict[str, Any]:
    """Convert row to dict."""

    result: dict[str, Any] = {}

    for index, header in enumerate(headers):
        if not header:
            continue

        result[header] = row[index] if index < len(row) else None

    return result


def row_is_empty(
    row: tuple[Any, ...],
) -> bool:
    """Return whether row is empty."""

    return all(value is None or str(value).strip() == "" for value in row)


def normalize_text(
    value: object,
) -> str:
    """Normalize text."""

    return str(value or "").strip()


def get_text(
    row: dict[str, Any],
    key: str,
) -> str:
    """Get optional text field."""

    return normalize_text(row.get(key))


def get_required_text(
    row: dict[str, Any],
    key: str,
    row_index: int,
) -> str:
    """Get required text field."""

    value = get_text(row, key)

    if not value:
        raise ValueError(f"row {row_index}: required field is empty: {key}")

    return value


def parse_semicolon_tuple(
    value: object,
) -> tuple[str, ...]:
    """Parse semicolon separated values."""

    text = normalize_text(value)

    if not text:
        return ()

    return tuple(
        item.strip()
        for item in re.split(r"[;；]", text)
        if item.strip()
    )


def parse_bool(
    value: object,
) -> bool:
    """Parse bool-like value."""

    text = normalize_text(value).lower()

    if text in {"true", "1", "yes", "y"}:
        return True

    if text in {"false", "0", "no", "n", ""}:
        return False

    raise ValueError(f"invalid bool value: {value!r}")

class LogisticsChunkBuilder:
    def load(self, logistics_file, collection_name="logistics_kb_v1"):
        return build_logistics_kb_chunks(
            logistics_file=logistics_file,
            collection_name=collection_name,
        )
