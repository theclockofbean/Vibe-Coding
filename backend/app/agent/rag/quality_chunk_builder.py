"""Build quality KB chunks from quality_questions.xlsx."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]

QUALITY_MODULE: Final[str] = "quality"
DEFAULT_LANGUAGE: Final[str] = "zh"
DEFAULT_SOURCE_TYPE: Final[str] = "qa_pair"
DEFAULT_SOURCE_NAME: Final[str] = "quality_questions.xlsx"

LOW_RISK_FLAGS: Final[set[str]] = {
    "",
}

MEDIUM_RISK_FLAGS: Final[set[str]] = {
    "unsupported_claim",
}

HIGH_RISK_FLAGS: Final[set[str]] = {
    "safety_claim",
    "quality_guarantee",
    "durability_claim",
    "certification_claim",
    "inspection_report_required",
    "aftersale_commitment",
    "compensation_claim",
    "installation_advice",
    "vehicle_fitment_unverified",
    "data_conflict",
    "business_policy_confirmation",
}


@dataclass(frozen=True)
class QualityQARecord:
    """Normalized quality QA record from Excel."""

    qa_id: str
    source_group_id: str
    question_raw: str
    question_normalized: str
    answer_raw: str
    answer_standard: str
    primary_intent: str
    intent_subtype: str
    secondary_intents: list[str] = field(default_factory=list)
    related_sku_ids: list[str] = field(default_factory=list)
    risk_flags: list[str] = field(default_factory=list)
    handoff_required: bool = False
    verification_status: str = "pending"
    expected_source: str = ""
    notes: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)
    excel_row: int = 0


@dataclass(frozen=True)
class QualityKBChunk:
    """Quality KB chunk prepared for knowledge_chunks and Qdrant."""

    chunk_id: str
    module: str
    source_type: str
    source_name: str
    source_uri: str
    doc_id: str
    doc_title: str
    chunk_index: int
    sku_scope: list[str]
    intent_scope: list[str]
    content: str
    content_hash: str
    summary: str
    language: str
    risk_level: str
    is_active: bool
    is_verified: bool
    allow_answer_reference: bool
    allow_commitment_reference: bool
    metadata: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        """Return serializable dict."""

        return {
            "chunk_id": self.chunk_id,
            "module": self.module,
            "source_type": self.source_type,
            "source_name": self.source_name,
            "source_uri": self.source_uri,
            "doc_id": self.doc_id,
            "doc_title": self.doc_title,
            "chunk_index": self.chunk_index,
            "sku_scope": self.sku_scope,
            "intent_scope": self.intent_scope,
            "content": self.content,
            "content_hash": self.content_hash,
            "summary": self.summary,
            "language": self.language,
            "risk_level": self.risk_level,
            "is_active": self.is_active,
            "is_verified": self.is_verified,
            "allow_answer_reference": self.allow_answer_reference,
            "allow_commitment_reference": self.allow_commitment_reference,
            "metadata": self.metadata,
        }


def load_quality_qa_records(
    *,
    workbook_path: Path,
    sheet_name: str = "qa_pairs",
) -> list[QualityQARecord]:
    """Load normalized quality QA records from workbook."""

    if not workbook_path.exists():
        raise FileNotFoundError(f"quality workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"sheet not found: {sheet_name}; available={workbook.sheetnames}"
        )

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise RuntimeError(f"sheet is empty: {sheet_name}")

    headers = [
        text(value)
        for value in rows[0]
    ]

    records: list[QualityQARecord] = []

    for excel_row, values in enumerate(rows[1:], start=2):
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }

        if all(not text(value) for value in row.values()):
            continue

        record = QualityQARecord(
            qa_id=text(row.get("qa_id")),
            source_group_id=text(row.get("source_group_id")),
            question_raw=text(row.get("question_raw")),
            question_normalized=text(row.get("question_normalized")),
            answer_raw=text(row.get("answer_raw")),
            answer_standard=text(row.get("answer_standard")),
            primary_intent=text(row.get("primary_intent")),
            intent_subtype=text(row.get("intent_subtype")),
            secondary_intents=split_semicolon(row.get("secondary_intents")),
            related_sku_ids=split_semicolon(row.get("related_sku_ids")),
            risk_flags=split_semicolon(row.get("risk_flags")),
            handoff_required=text(row.get("handoff_required")).lower() == "true",
            verification_status=text(row.get("verification_status")).lower()
            or "pending",
            expected_source=text(row.get("expected_source")),
            notes=text(row.get("notes")),
            metadata=parse_metadata(row.get("metadata")),
            excel_row=excel_row,
        )

        records.append(record)

    return records


def build_quality_kb_chunks(
    *,
    records: list[QualityQARecord],
    source_uri: str,
) -> list[QualityKBChunk]:
    """Build quality KB chunks from records."""

    chunks: list[QualityKBChunk] = []

    for index, record in enumerate(records):
        chunk = build_quality_kb_chunk(
            record=record,
            source_uri=source_uri,
            chunk_index=index,
        )
        chunks.append(chunk)

    return chunks


def build_quality_kb_chunk(
    *,
    record: QualityQARecord,
    source_uri: str,
    chunk_index: int,
) -> QualityKBChunk:
    """Build one quality KB chunk."""

    content = build_quality_content(record)
    content_hash = hashlib.sha256(content.encode("utf-8")).hexdigest()
    risk_level = infer_risk_level(record)
    is_verified = record.verification_status == "verified"

    return QualityKBChunk(
        chunk_id=f"quality_qa_{record.qa_id.lower()}",
        module=QUALITY_MODULE,
        source_type=DEFAULT_SOURCE_TYPE,
        source_name=DEFAULT_SOURCE_NAME,
        source_uri=source_uri,
        doc_id=record.qa_id,
        doc_title=f"品质问答 {record.qa_id}｜{record.intent_subtype}",
        chunk_index=chunk_index,
        sku_scope=record.related_sku_ids,
        intent_scope=build_intent_scope(record),
        content=content,
        content_hash=content_hash,
        summary=build_summary(record),
        language=DEFAULT_LANGUAGE,
        risk_level=risk_level,
        is_active=record.verification_status != "rejected",
        is_verified=is_verified,
        allow_answer_reference=record.verification_status != "rejected",
        allow_commitment_reference=False,
        metadata=build_chunk_metadata(record),
    )


def build_quality_content(
    record: QualityQARecord,
) -> str:
    """Build final chunk content."""

    sku_scope = ";".join(record.related_sku_ids) or "未限定"
    risk_flags = ";".join(record.risk_flags) or "none"
    secondary_intents = ";".join(record.secondary_intents) or "none"

    return "\n".join(
        [
            "【问题】",
            record.question_normalized,
            "",
            "【标准回答】",
            record.answer_standard,
            "",
            "【适用范围】",
            f"SKU: {sku_scope}",
            f"primary_intent: {record.primary_intent}",
            f"intent_subtype: {record.intent_subtype}",
            f"secondary_intents: {secondary_intents}",
            "",
            "【风险边界】",
            f"risk_flags: {risk_flags}",
            f"handoff_required: {str(record.handoff_required).lower()}",
            f"verification_status: {record.verification_status}",
            "",
            "【使用限制】",
            "该内容只能作为品质类解释与资料补充。",
            "不得据此生成绝对质量、耐久、售后、赔付或适配承诺。",
            "如问题涉及检测记录、认证、安全、投诉、赔付或售后，应转人工确认。",
        ]
    )


def infer_risk_level(
    record: QualityQARecord,
) -> str:
    """Infer risk level."""

    risk_flags = set(record.risk_flags)

    if risk_flags & HIGH_RISK_FLAGS:
        return "high"

    if record.handoff_required:
        return "high"

    if risk_flags & MEDIUM_RISK_FLAGS:
        return "medium"

    return "low"


def build_summary(
    record: QualityQARecord,
) -> str:
    """Build short summary."""

    answer = record.answer_standard.replace("\n", " ").strip()

    if len(answer) > 120:
        answer = answer[:117] + "..."

    return f"{record.question_normalized}｜{answer}"


def build_intent_scope(
    record: QualityQARecord,
) -> list[str]:
    """Build intent scope."""

    scope = ["quality"]

    if record.intent_subtype:
        scope.append(record.intent_subtype)

    for secondary_intent in record.secondary_intents:
        if secondary_intent and secondary_intent not in scope:
            scope.append(secondary_intent)

    return scope


def build_chunk_metadata(
    record: QualityQARecord,
) -> dict[str, Any]:
    """Build chunk metadata."""

    return {
        "qa_id": record.qa_id,
        "source_group_id": record.source_group_id,
        "question_raw": record.question_raw,
        "answer_raw": record.answer_raw,
        "expected_source": record.expected_source,
        "notes": record.notes,
        "excel_row": record.excel_row,
        "record_metadata": record.metadata,
    }


def parse_metadata(
    value: object,
) -> dict[str, Any]:
    """Parse metadata cell."""

    raw_text = text(value)

    if not raw_text:
        return {}

    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return {
            "raw_metadata": raw_text,
        }

    if isinstance(parsed, dict):
        return {
            str(key): item_value
            for key, item_value in parsed.items()
        }

    return {
        "raw_metadata": parsed,
    }


def split_semicolon(
    value: object,
) -> list[str]:
    """Split semicolon separated text."""

    raw_text = text(value)

    if not raw_text:
        return []

    return [
        item.strip()
        for item in raw_text.split(";")
        if item.strip()
    ]


def text(
    value: object,
) -> str:
    """Return stripped text."""

    if value is None:
        return ""

    return str(value).strip()

class QualityChunkBuilder:
    def load(self, quality_file, collection_name="quality_kb_v1"):
        return build_quality_kb_chunks(
            quality_file=quality_file,
            collection_name=collection_name,
        )