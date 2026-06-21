from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from app.agent.rag.chunk_schema import DocumentChunk


DEFAULT_CHUNK_SIZE = 1000
DEFAULT_CHUNK_OVERLAP = 120


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").replace("\r\n", "\n").replace("\r", "\n").strip()


def _stable_id(*parts: Any) -> str:
    raw = "||".join(_normalize_text(part) for part in parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _split_text(text: str, chunk_size: int, chunk_overlap: int) -> List[str]:
    text = _normalize_text(text)

    if not text:
        return []

    if chunk_size <= 0:
        raise ValueError("chunk_size must be greater than 0")

    if chunk_overlap < 0:
        raise ValueError("chunk_overlap must be greater than or equal to 0")

    if chunk_overlap >= chunk_size:
        raise ValueError("chunk_overlap must be smaller than chunk_size")

    chunks: List[str] = []
    start = 0

    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end].strip()

        if chunk:
            chunks.append(chunk)

        if end >= len(text):
            break

        start = end - chunk_overlap

    return chunks


def _dict_to_text(row: Dict[str, Any]) -> str:
    lines: List[str] = []

    for key, value in row.items():
        value_text = _normalize_text(value)
        if value_text:
            lines.append(f"{key}: {value_text}")

    return "\n".join(lines).strip()


class DocumentLoader:
    """
    RAG 文档加载器。

    职责：
    1. 从 txt / md / json / jsonl / xlsx 加载原始知识
    2. 转换为统一 DocumentChunk
    3. 为后续 embedding、Qdrant 入库、citation 提供稳定字段
    """

    def __init__(
        self,
        chunk_size: int = DEFAULT_CHUNK_SIZE,
        chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
    ) -> None:
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    def load_text(
        self,
        text: str,
        *,
        domain: str,
        source: str,
        source_type: str = "doc",
        document_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> List[DocumentChunk]:
        document_id = document_id or _stable_id(source, domain)
        metadata = metadata or {}

        text_chunks = _split_text(
            text=text,
            chunk_size=self.chunk_size,
            chunk_overlap=self.chunk_overlap,
        )

        chunks: List[DocumentChunk] = []

        for index, chunk_text in enumerate(text_chunks):
            chunk_id = _stable_id(document_id, index, chunk_text)

            chunks.append(
                DocumentChunk(
                    chunk_id=chunk_id,
                    document_id=document_id,
                    domain=domain,
                    text=chunk_text,
                    source=source,
                    source_type=source_type,
                    metadata={
                        **metadata,
                        "chunk_index": index,
                    },
                )
            )

        return chunks

    def load_text_file(
        self,
        file_path: str | Path,
        *,
        domain: str,
        source_type: str = "doc",
        encoding: str = "utf-8",
        metadata: Optional[Dict[str, Any]] = None,
    ) -> List[DocumentChunk]:
        path = Path(file_path)
        text = path.read_text(encoding=encoding)

        return self.load_text(
            text,
            domain=domain,
            source=str(path),
            source_type=source_type,
            document_id=_stable_id(str(path), domain),
            metadata={
                **(metadata or {}),
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "source_path": str(path),
            },
        )

    def load_json_file(
        self,
        file_path: str | Path,
        *,
        domain: str,
        source_type: str = "faq",
        encoding: str = "utf-8",
        metadata: Optional[Dict[str, Any]] = None,
    ) -> List[DocumentChunk]:
        path = Path(file_path)
        data = json.loads(path.read_text(encoding=encoding))

        if isinstance(data, dict):
            rows: Iterable[Any] = data.get("items") or data.get("data") or [data]
        elif isinstance(data, list):
            rows = data
        else:
            raise ValueError(f"Unsupported json root type: {type(data)}")

        return self.load_records(
            rows,
            domain=domain,
            source=str(path),
            source_type=source_type,
            document_id=_stable_id(str(path), domain),
            metadata={
                **(metadata or {}),
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "source_path": str(path),
            },
        )

    def load_jsonl_file(
        self,
        file_path: str | Path,
        *,
        domain: str,
        source_type: str = "faq",
        encoding: str = "utf-8",
        metadata: Optional[Dict[str, Any]] = None,
    ) -> List[DocumentChunk]:
        path = Path(file_path)
        rows: List[Dict[str, Any]] = []

        for line_number, line in enumerate(path.read_text(encoding=encoding).splitlines(), start=1):
            line = line.strip()
            if not line:
                continue

            item = json.loads(line)

            if not isinstance(item, dict):
                raise ValueError(f"JSONL line {line_number} must be an object")

            rows.append(
                {
                    **item,
                    "_line_number": line_number,
                }
            )

        return self.load_records(
            rows,
            domain=domain,
            source=str(path),
            source_type=source_type,
            document_id=_stable_id(str(path), domain),
            metadata={
                **(metadata or {}),
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "source_path": str(path),
            },
        )

    def load_records(
        self,
        rows: Iterable[Any],
        *,
        domain: str,
        source: str,
        source_type: str,
        document_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> List[DocumentChunk]:
        document_id = document_id or _stable_id(source, domain)
        metadata = metadata or {}

        chunks: List[DocumentChunk] = []

        for row_index, row in enumerate(rows):
            if isinstance(row, dict):
                from app.agent.rag.embedding_schema import build_embedding_text
                text = build_embedding_text(row)
                row_metadata = row
            else:
                text = _normalize_text(row)
                row_metadata = {"value": row}

            if not text:
                continue

            chunk_id = _stable_id(document_id, row_index, text)

            chunks.append(
                DocumentChunk(
                    chunk_id=chunk_id,
                    document_id=document_id,
                    domain=domain,
                    text=text,
                    source=source,
                    source_type=source_type,
                    metadata={
                        **metadata,
                        "row_index": row_index,
                        "row": row_metadata,
                    },
                )
            )

        return chunks

    def load_xlsx_file(
        self,
        file_path: str | Path,
        *,
        domain: str,
        source_type: str = "sku",
        sheet_name: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> List[DocumentChunk]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to load xlsx files. Install it with: pip install openpyxl"
            ) from exc

        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=True)

        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows_iter = worksheet.iter_rows(values_only=True)

        try:
            headers = next(rows_iter)
        except StopIteration:
            return []

        header_names = [
            _normalize_text(header) or f"column_{index + 1}"
            for index, header in enumerate(headers)
        ]

        rows: List[Dict[str, Any]] = []

        for excel_row_number, values in enumerate(rows_iter, start=2):
            row = {
                header_names[index]: values[index] if index < len(values) else None
                for index in range(len(header_names))
            }

            if any(_normalize_text(value) for value in row.values()):
                rows.append(
                    {
                        **row,
                        "_excel_row_number": excel_row_number,
                    }
                )

        return self.load_records(
            rows,
            domain=domain,
            source=str(path),
            source_type=source_type,
            document_id=_stable_id(str(path), domain, worksheet.title),
            metadata={
                **(metadata or {}),
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "source_path": str(path),
                "sheet_name": worksheet.title,
            },
        )


