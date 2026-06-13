"""Build Spec KB chunks from spec_questions.xlsx."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook  # type: ignore[import-untyped]


DEFAULT_COLLECTION_NAME: Final[str] = "spec_kb_v1"
DEFAULT_SOURCE_NAME: Final[str] = "spec_questions.xlsx"
DEFAULT_SHEET_NAME: Final[str] = "qa_pairs"


@dataclass(frozen=True)
class SpecKBChunk:
    """Spec KB chunk."""

    chunk_id: str
    doc_id: str
    doc_title: str
    content: str
    summary: str
    source_row_index: int
    collection_name: str
    module: str
    source_type: str
    source_name: str
    qa_id: str
    source_group_id: str
    primary_intent: str
    secondary_intents: list[str]
    intent_subtype: str
    question_raw: str
    question_normalized: str
    answer_raw: str
    answer_standard: str
    related_sku_ids: list[str]
    required_fields: list[str]
    answer_source: str
    handoff_required: bool
    risk_flags: list[str]
    verification_status: str
    review_notes: str
    allow_answer_reference: bool
    allow_commitment_reference: bool
    risk_level: str

    def to_qdrant_payload(self) -> dict[str, Any]:
        """Return Qdrant payload."""

        return {
            "chunk_id": self.chunk_id,
            "doc_id": self.doc_id,
            "doc_title": self.doc_title,
            "content": self.content,
            "summary": self.summary,
            "source_row_index": self.source_row_index,
            "collection_name": self.collection_name,
            "module": self.module,
            "source_type": self.source_type,
            "source_name": self.source_name,
            "qa_id": self.qa_id,
            "source_group_id": self.source_group_id,
            "primary_intent": self.primary_intent,
            "secondary_intents": self.secondary_intents,
            "intent_subtype": self.intent_subtype,
            "question_raw": self.question_raw,
            "question_normalized": self.question_normalized,
            "answer_raw": self.answer_raw,
            "answer_standard": self.answer_standard,
            "related_sku_ids": self.related_sku_ids,
            "required_fields": self.required_fields,
            "answer_source": self.answer_source,
            "handoff_required": self.handoff_required,
            "risk_flags": self.risk_flags,
            "verification_status": self.verification_status,
            "review_notes": self.review_notes,
            "allow_answer_reference": self.allow_answer_reference,
            "allow_commitment_reference": self.allow_commitment_reference,
            "risk_level": self.risk_level,
            "is_verified": self.verification_status.lower() in {"verified", "approved"},
        }


def build_spec_kb_chunks_from_excel(
    *,
    spec_file: Path,
    collection_name: str = DEFAULT_COLLECTION_NAME,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[SpecKBChunk]:
    """Build Spec KB chunks from Excel file."""

    workbook = load_workbook(spec_file, read_only=True, data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"missing sheet: {sheet_name}")

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"empty sheet: {sheet_name}")

    headers = [
        str(value).strip() if value is not None else ""
        for value in rows[0]
    ]

    chunks: list[SpecKBChunk] = []

    for source_row_index, row in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue

        record = dict(zip(headers, row, strict=False))
        chunks.append(
            build_spec_kb_chunk(
                record=record,
                source_row_index=source_row_index,
                collection_name=collection_name,
            )
        )

    return chunks


def build_spec_kb_chunk(
    *,
    record: dict[str, Any],
    source_row_index: int,
    collection_name: str = DEFAULT_COLLECTION_NAME,
) -> SpecKBChunk:
    """Build one Spec KB chunk."""

    qa_id = read_text(record, "qa_id")
    source_group_id = read_text(record, "source_group_id")
    primary_intent = read_text(record, "primary_intent")
    secondary_intents = split_multi_value(read_text(record, "secondary_intents"))
    intent_subtype = read_text(record, "intent_subtype")
    question_raw = read_text(record, "question_raw")
    question_normalized = read_text(record, "question_normalized")
    answer_raw = read_text(record, "answer_raw")
    answer_standard = read_text(record, "answer_standard")
    related_sku_ids = split_multi_value(read_text(record, "related_sku_ids"))
    required_fields = split_multi_value(read_text(record, "required_fields"))
    answer_source = read_text(record, "answer_source")
    handoff_required = parse_bool(read_text(record, "handoff_required"))
    risk_flags = split_multi_value(read_text(record, "risk_flags"))
    verification_status = read_text(record, "verification_status")
    review_notes = read_text(record, "review_notes")

    if not qa_id:
        raise ValueError(f"row {source_row_index}: empty qa_id")

    if primary_intent != "spec":
        raise ValueError(
            f"row {source_row_index}: primary_intent must be spec, got {primary_intent}"
        )

    title_question = question_normalized or question_raw
    doc_title = f"{qa_id}｜规格问答｜{intent_subtype}｜{title_question}"
    risk_level = infer_risk_level(
        handoff_required=handoff_required,
        risk_flags=risk_flags,
    )

    content = build_chunk_content(
        qa_id=qa_id,
        source_group_id=source_group_id,
        primary_intent=primary_intent,
        secondary_intents=secondary_intents,
        intent_subtype=intent_subtype,
        question_raw=question_raw,
        question_normalized=question_normalized,
        answer_raw=answer_raw,
        answer_standard=answer_standard,
        related_sku_ids=related_sku_ids,
        required_fields=required_fields,
        answer_source=answer_source,
        handoff_required=handoff_required,
        risk_flags=risk_flags,
        verification_status=verification_status,
        review_notes=review_notes,
    )

    return SpecKBChunk(
        chunk_id=f"spec_qa_{qa_id.lower()}",
        doc_id=f"spec_qa_{qa_id.lower()}",
        doc_title=doc_title,
        content=content,
        summary=f"{qa_id}｜规格问答｜{intent_subtype}｜{title_question}",
        source_row_index=source_row_index,
        collection_name=collection_name,
        module="spec",
        source_type="qa_pair",
        source_name=DEFAULT_SOURCE_NAME,
        qa_id=qa_id,
        source_group_id=source_group_id,
        primary_intent=primary_intent,
        secondary_intents=secondary_intents,
        intent_subtype=intent_subtype,
        question_raw=question_raw,
        question_normalized=question_normalized,
        answer_raw=answer_raw,
        answer_standard=answer_standard,
        related_sku_ids=related_sku_ids,
        required_fields=required_fields,
        answer_source=answer_source,
        handoff_required=handoff_required,
        risk_flags=risk_flags,
        verification_status=verification_status,
        review_notes=review_notes,
        allow_answer_reference=True,
        allow_commitment_reference=False,
        risk_level=risk_level,
    )


def build_chunk_content(
    *,
    qa_id: str,
    source_group_id: str,
    primary_intent: str,
    secondary_intents: list[str],
    intent_subtype: str,
    question_raw: str,
    question_normalized: str,
    answer_raw: str,
    answer_standard: str,
    related_sku_ids: list[str],
    required_fields: list[str],
    answer_source: str,
    handoff_required: bool,
    risk_flags: list[str],
    verification_status: str,
    review_notes: str,
) -> str:
    """Build chunk content for embedding."""

    return "\n".join(
        [
            f"QA编号：{qa_id}",
            f"来源组：{source_group_id}",
            f"主意图：{primary_intent}",
            f"次级意图：{format_list(secondary_intents)}",
            f"规格子类型：{intent_subtype}",
            f"原始问题：{question_raw}",
            f"规范问题：{question_normalized}",
            f"原始回答：{answer_raw}",
            f"标准回答：{answer_standard}",
            f"关联SKU：{format_list(related_sku_ids)}",
            f"必需字段：{format_list(required_fields)}",
            f"回答依据：{answer_source}",
            f"是否需人工确认：{handoff_required}",
            f"风险标记：{format_list(risk_flags)}",
            f"审核状态：{verification_status}",
            f"备注：{review_notes}",
            "边界规则：规格类问题必须以结构化 SKU 主数据和已审核问答为准。",
            "边界规则：不得承诺通用适配、万能适配或未经验证的兼容关系。",
        ]
    )


def infer_risk_level(
    *,
    handoff_required: bool,
    risk_flags: list[str],
) -> str:
    """Infer risk level."""

    if handoff_required or risk_flags:
        return "high"

    return "medium"


def read_text(
    record: dict[str, Any],
    key: str,
) -> str:
    """Read cell text."""

    value = record.get(key)

    if value is None:
        return ""

    return str(value).strip()


def split_multi_value(
    value: str,
) -> list[str]:
    """Split multi-value cell."""

    if not value:
        return []

    normalized = (
        value.replace("，", ",")
        .replace("；", ",")
        .replace(";", ",")
        .replace("|", ",")
        .replace("/", ",")
    )

    return [
        item.strip()
        for item in normalized.split(",")
        if item.strip()
    ]


def parse_bool(
    value: str,
) -> bool:
    """Parse boolean-ish value."""

    normalized = value.strip().lower()

    return normalized in {"1", "true", "yes", "y", "是", "需", "需要"}


def format_list(
    values: list[str],
) -> str:
    """Format list."""

    if not values:
        return "无"

    return "、".join(values)